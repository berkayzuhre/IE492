#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Created by Tunc Ali Kütükcüoglu (in 2019)
# see: https://software.tuncalik.com/travel-and-assignment-planning-software-in-python/4812
# Copyrights: Tunc Ali Kütükcüoglu (senior data analyst & developer)
"""
Tour Planning: Search for all possible paths (tours) 
according to given search and filter criteria.

Timetable data in PostgreSQL database (table named timetable)
"""
import psycopg2
import sys, os
import time
import math
from timeit import default_timer
import timeit
from datetime import date 
from datetime import datetime
from datetime import timedelta
from sklearn.cluster import KMeans
from matplotlib import pyplot as plt
from openpyxl import Workbook
import openpyxl
import csv
import pandas as pd

import calendar
import itertools as it

from BU2019_CentralParameters import *
from BU2019_BasicFunctionsLib import *
from GlobalVariables import*

# **************************************************************************************
# Read some global parameters from database tables
# **************************************************************************************


# **************************************************************************************
# Defining search conditions for routes (tours) with parameters
# **************************************************************************************

class Cond:
	"""
	Class with constants for defining conditions with parameters
	"""
	# class variables
	SelectedRoutes = []
	TerminationReasonsDic = {}					# search report
	RouteCountPerRouteLength = {}				# number of routes per route length (#connections) after route selection, before final filtering
	
	RouteCountAfterConnectionSelection = 0 		# number of routes after connection selection (before route selection)
	RouteCountAfterRouteSelection = 0 			# number of routes after route selection (before final filtering of routes)
	RouteCountAfterRouteFiltering = 0 			# remaining final route count after route filtering

	SearchStartTime = None 
	MaxDurationSinceArrivalToFirstStation = 0 
	MaxDurationSinceDepartureFromFirstStation = 0 
	RouteSearchReportCounter = 1 				# used for reporting frequency

	IfTestRouteSearch = False 
	TestWaitingTime = None

	MinimumDuration = None 			# min duration of route in minutes
	EarliestArrival = None			# earliest arrival at (in total minutes like 560)
	MinLineChangeCount = None

	CoveredLineKeys = {}			# How many times a LineKey is covered so far, during tour search
	CheckFunctionPerCondition = {}	# boolean check function of condition
	
	# class constants
	StartTimeAndDuration = 1
	StartTimeAndDuration_explain = """
		Set parameters for start time and duration of route. 
		This is a mandatory condition.
		Parameters: (StartHour, StartMin, EarliestArrivalIn, LatestArrivalIn)
		LatestArrivalIn: Latest Arrival in ... minutes (interval)
		EarliestArrivalIn: Earliest Arrival in ... minutes (interval)
		Example: Cond.StartTimeAndDuration: (15, 45, 3*60, 6*60)
		"""

	SelectWeekDays = 2
	SelectWeekDays_explain = """
		Select week days (Mon-Sun) for all connections of the route.
		All connections of the route must be available in one of the 
		given week days.
		Parameters: WeekDayList
		Example: Cond.SelectWeekDays: ([1,2,3,4,5],)
		1 means Monday, 7 is Sunday
		"""

	MaxWaitingTimeAtStation = 3
	MaxWaitingTimeAtStation_explain = """
		Maximum waiting time at a station (in minutes) before taking 
		the next connection. This is a mandatory condition.
		Parameters: MaxWaitingTime in minutes
		Example: Cond.MaxWaitingTimeAtStation: (30,)
		"""

	MaxStationCount = 4
	MaxStationCount_explain = """
		Maximum number of stations a route can contain. 
		Example: Cond.MaxStationCount: (10,)
		"""

	MinStationCount = 5
	MinStationCount_explain = """
		Minimum number of stations a route can contain. 
		Example: Cond.MinStationCount: (4,)
		"""
	DefaultTimeForLineChange = 6 
	DefaultTimeForLineChange_explain = """
		Minimum default time in minutes required for changing the line. 
		This time is only valid if the line-change time is not given 
		specifically for a station (e.g. in hafas text data UMSTEIGB).
		This is a mandatory condition.
		Parameters: ChangeTime in minutes
		Example: Cond.DefaultTimeForLineChange: (5,)
		"""

	# alias values (legacy)
	TimeForLineChange = DefaultTimeForLineChange
	TimeForLineChange_explain = DefaultTimeForLineChange_explain

	VisitAStationOnlyOnce = 7
	VisitAStationOnlyOnce_explain = """
		Simple path condition: Visit a station only once, 
		possibly excluding the first and last stations for a round trip.
		Example: Cond.VisitAStationOnlyOnce: True
		No parameter is required.
		"""

	IncludeListedGattungsOnly = 8
	IncludeListedGattungsOnly_explain = """
		Select lines only from the listed gattungs (line category).
		Parameters: GattungList
		Example: Cond.IncludeListedGattungsOnly: (['S','R','BUS'],)
		"""

	VisitStations = 11 
	VisitStations_explain = """
		Visit stations given in the list of stations (StationList).
		Parameters: StationList, IncludeOption
		IncludeOption: (how to include stations in StationList)
		INCLUDE_ALL_AND_ONLY 	# visit all and only listed stations
		INCLUDE_ALL  			# visit all listed stations (and possibly more)
		INCLUDE_ONLY 			# visit only (some of) listed stations
		Example: Cond.VisitStations: ([8503000, 8503006, 8503129], INCLUDE_ALL)
		"""

	ConnectionsAreAvailableOnAllListedDays = 12 
	ConnectionsAreAvailableOnAllListedDays_explain = """
		All connections of the route must be available on all listed days in DayList.
		Parameters: DayOrdList (list of ordinal dates)
		Example: Cond.ConnectionsAreAvailableOnAllListedDays: (DayOrdList,)
		"""

	ReturnFromCurrentStation = 13
	ReturnFromCurrentStation_explain= """
		Checks if we can return to our starting point from current station in the tour.
		No parameter is required.
		"""

	StartAndEndStations = 17
	StartAndEndStations_explain = """
		Set start and end stations of the route.
		Parameters: StartStation, EndStation
		Example: Cond.StartAndEndStations: (8503003, 8503000)
		"""

	MaxNumberOfLineChanges = 20 
	MaxNumberOfLineChanges_explain = """
		Max number of line changes (Max Anzahl Umsteigen) in a tour. 
		0 as parameter value means no line change is permitted.
		A station passage alone is not counted as an additional line change.
		Example: Cond.MaxNumberOfLineChanges: 1
		"""

	ExcludeListedGattungs = 22
	ExcludeListedGattungs_explain = """
		Exclude listed gattungs (line category) from selected connections.
		Parameters: ExcludeGattungList
		Example: Cond.ExcludeListedGattungs: (['S','R','BUS'],)
		"""

	IncludeListedManagementsOnly = 23
	IncludeListedManagementsOnly_explain = """
		Select connections only from the listed managements.
		Parameters: ManagementList
		Example: Cond.IncludeListedManagementssOnly: ([11,23,111],)
		"""

	MaxSearchTimeInSeconds = 29 
	MaxSearchTimeInSeconds_explain = """
		Limit the maximum search execution time (in seconds).
		Return all the solutions found so far, if execution time was exceeded.
		Parameter: MaxSearchTimeInSeconds
		Example: 
		Cond.MaxSearchTimeInSeconds: (30,)
		"""

	ReportDuringRouteSearch = 30
	ReportDuringRouteSearch_explain = """
		Report during route search: How many routes were found so far, etc.
		Parameter: RouteSearchReportingIntervalInSeconds
		Example: 
		Cond.ReportDuringRouteSearch: (10,)
		"""

	TestRunDuringRouteSearch = 31 
	TestRunDuringRouteSearch_explain = """
		Make a test run of rute search, by displaying intermediate results 
		like station-path, ConnectionInfoList, termination reasons etc.
		Parameter: WaitingTimeInSeconds (waiting time between each iteration)
		Example: Cond.TestRunDuringRouteSearch: (0.5,)
		"""

	SearchRoutesForEarliestArrival = 35 
	SearchRoutesForEarliestArrival_explain = """
		Find routes that reach end station at the earliest possible time.
		
		If CheckMinLineChange is set to True, minimum number of line changes 
		becomes the primary selection criterion before earliest arrival.
		
		Parameters: 
			1) CheckMinLineChange
		Example: Cond.SearchRoutesForEarliestArrival: (False,)
		"""

	MaxTripDurationSinceDepartureFromTheFirstStation = 43 
	MaxTripDurationSinceDepartureFromTheFirstStation_explain = """
		Maximum trip duration in minutes since departure from the 
		first station. Waiting time at the first station is not included.
		Parameter: MaxTripDurInMinutes
		Example: Cond.MaxTripDuration: (60,)
		"""

	VisitStationsInGivenOrder = 59 
	VisitStationsInGivenOrder_explain = """
		Visit all stations in the list, exactly in given order (sequence).

		That is, in a station list like [s1, s2, s3,..] s2 cannot be visited befire s1.
		There can be many other stations between subsequent stations in the list.

		Parameters: OrderedStationList
			like [8503000, 8503574, 8502119, 8572373, 8507000)]
			meaning: Zürich-->Brugg-->Lenzburg-->Solothurn-->Bern

		Example: 
			Cond.VisitStationsInGivenOrder: (OrderedStationList,)
		"""

	ReachingOneOfTheStations = 1 
	ReachingOneOfTheLineIDs = 2 

	SuccessfulTerminationBy = 36
	SuccessfulTerminationBy_explain = """
		Terminate a path with success if one of the stations or LineIDs in 
		the given list is reached. 

		Overwrites the standard "end station" condition for terminating a 
		path successfully.

		An alternative way of terminating a path successfully, other than 
		the standard end station. 

		Parameters:
		1) TerminationType: Cond.ReachingOneOfTheStations or Cond.ReachingOneOfTheLineIDs
		2) List of stations or LineIDs depending on TerminationType 
		Example: Cond.SuccessfulTerminationBy: (Cond.ReachingOneOfTheStations, StationList)
		"""

	VisitAConnectionOnlyOnce = 100
	VisitAConnectionOnlyOnce_explain = """
		Visit a connection (line segment (edge) connecting two neigbour stations) only once,
		in both directions. That is, if the path has already visited the connection (st1,st2),
		it cannot visit (st1,st2) or (st2,st1) again.

		Parameters: none
		Example: Cond.VisitAConnectionOnlyOnce: ()
		"""

	@classmethod
	def ResetClassVariables(cls):
		"""
		Reset class variables, and return status variables.

		Returns:
		(StatusReport, TerminationReasons)
		"""
		# get values to return
		StatusReport = {}
		StatusReport['1) RouteCountAfterConnectionSelection'] = cls.RouteCountAfterConnectionSelection
		StatusReport['2) RouteCountAfterRouteSelection'] = cls.RouteCountAfterRouteSelection
		StatusReport['3) RouteCountAfterRouteFiltering'] = cls.RouteCountAfterRouteFiltering

		# see: http://stackoverflow.com/questions/2465921/how-to-copy-a-dictionary-and-only-edit-the-copy
		TerminationReasons = cls.TerminationReasonsDic.copy()

		# class variables
		cls.SelectedRoutes = []
		cls.TerminationReasonsDic = {}					
		cls.RouteCountAfterConnectionSelection = 0 		
		cls.RouteCountAfterRouteSelection = 0 
		cls.RouteCountAfterRouteFiltering = 0			
		
		cls.SearchStartTime = None 
		cls.MaxDurationSinceArrivalToFirstStation = 0 
		cls.MaxDurationSinceDepartureFromFirstStation = 0 
		cls.RouteSearchReportCounter = 1 		
		cls.IfTestRouteSearch = False 
		cls.TestWaitingTime = None	
		
		cls.MinimumDuration = None 			
		cls.EarliestArrival = None
		cls.MinLineChangeCount = None
		cls.CoveredLineKeys = {}

		return (StatusReport, TerminationReasons)

	@classmethod
	def GenerateSQLConditions(cls, RouteConditions):
		"""
		Generate all SQL condition strings according to the conditions
		defined in dictionary RouteConditions.
		Returns an aggregate condition string (SQLcond)
		"""
		SQLconditions = []

		for cond in RouteConditions:
			parameters = RouteConditions[cond]

			# StartTimeAndDuration
			if cond == cls.StartTimeAndDuration:
				selparameters = (parameters[0],parameters[1],parameters[3])
				SQLconditions.append(CondStrRouteStartAndDuration(*selparameters))

			# SelectWeekDays
			elif cond == cls.SelectWeekDays:
				SQLconditions.append(CondStrSelectedWeekDays(*parameters))

			# IncludeListedManagementsOnly
			elif cond == cls.IncludeListedManagementsOnly:
				SQLconditions.append(CondStrSelectedVerwaltungs(*parameters))

			# IncludeListedGattungsOnly
			elif cond == cls.IncludeListedGattungsOnly:
				SQLconditions.append(CondStrSelectedGattung(*parameters))

			# ExcludeListedGattungs
			elif cond == cls.ExcludeListedGattungs:
				SQLconditions.append(CondStrExcludeGattungs(*parameters))

			# VisitStations
			elif cond == cls.VisitStations:
				StationList = parameters[0]
				IncludeOption = parameters[1]
				if IncludeOption in (INCLUDE_ALL_AND_ONLY, INCLUDE_ONLY):
					SQLconditions.append(CondStrSelectedStations(StationList))

			# ConnectionsAreAvailableOnAllListedDays
			elif cond == cls.ConnectionsAreAvailableOnAllListedDays:
				DayList = parameters[0]
				SQLconditions.append(CondStrSelectedTripDays(DayList))

		# test
		# SQLconditions.append(CondStrSelectedStations([8503000, 8503006, 8503129, 8503306, 8503147]))
		# SQLconditions.append(CondStrSelectedLines([(11,'S','8'),(11,'S','11'),(11,'S','12'),(11,'S','15')]))

		return ' AND '.join(SQLconditions)

	@classmethod
	def CheckIfConnectionShouldBeSelected(cls, ConnectionInfo, PathInfo, EndStation, RouteConditions,EarliestArrival_StationScoring):
		"""
		Determine whether the given connection should be selected 
		according to the conditions defined in dictionary RouteConditions.
		Returns:
		1: True (check next condition)
		2: False (continue to the next ConnectionInfo in loop)
		3: None (return None)
		"""
		global TOTAL_PATH_DURATION
		# test 
		# print "RouteConditions:\n" + str(RouteConditions)
		# print "ConnectionInfo:" + str(ConnectionInfo)
		IfTest = cls.IfTestRouteSearch

		# set SearchStartTime
		if cls.SearchStartTime == None: 
			cls.SearchStartTime = default_timer()

		# MaxSearchTimeInSeconds
		if RouteConditions.has_key(cls.MaxSearchTimeInSeconds):
			cond = cls.MaxSearchTimeInSeconds
			parameters = RouteConditions[cond]
			MaxSearchTime = parameters[0]
			# terminate search algorithm if MaxSearchTime is exceeded
			if default_timer() - cls.SearchStartTime >= MaxSearchTime: 
				IncrementDicValue(cls.TerminationReasonsDic, 'MaxSearchTimeInSeconds')
				if IfTest: print "--------- MaxSearchTimeInSeconds exceeded ---------"
				return None

		# StartAndEndStations
		if not RouteConditions.has_key(cls.StartAndEndStations):
			raise Exception("Cond.StartAndEndStations is missing! This is a mandatory condition.")

		# MaxStationCount
		if RouteConditions.has_key(cls.MaxStationCount):
			cond = cls.MaxStationCount
			parameters = RouteConditions[cond]
			MaxStationCount = parameters[0]
			if not CheckMaxStationCount(ConnectionInfo, PathInfo, MaxStationCount):
				IncrementDicValue(cls.TerminationReasonsDic, 'MaxStationCount')
				if IfTest: print "--------- MaxStationCount exceeded ---------"
				return None
		
		# MaxWaitingTimeAtStation
		if not RouteConditions.has_key(cls.MaxWaitingTimeAtStation):
			raise Exception("Cond.MaxWaitingTimeAtStation is missing! This is a mandatory condition.")
		else:
			cond = cls.MaxWaitingTimeAtStation
			parameters = RouteConditions[cond]
			MaxWaitTime = parameters[0]
			if not CheckMaxWaitTimeAtStation(ConnectionInfo, PathInfo, MaxWaitTime):
				IncrementDicValue(cls.TerminationReasonsDic, 'MaxWaitingTimeAtStation')
				if IfTest: print "--------- MaxWaitingTimeAtStation exceeded ---------"
				return None	

		# MaxTripDurationSinceDepartureFromTheFirstStation
		if RouteConditions.has_key(cls.MaxTripDurationSinceDepartureFromTheFirstStation):
			cond = cls.MaxTripDurationSinceDepartureFromTheFirstStation
			parameters = RouteConditions[cond]
			MaxTripDuration = parameters[0]
			if not CheckTripDurationSinceDepartureFromFirstStation(ConnectionInfo, PathInfo, MaxTripDuration):
				IncrementDicValue(cls.TerminationReasonsDic, 'MaxTripDurationSinceDepartureFromTheFirstStation')
				if IfTest: print "--------- MaxTripDurationSinceDepartureFromTheFirstStation exceeded ---------"
				return None

		# StartTimeAndDuration
		if not RouteConditions.has_key(cls.StartTimeAndDuration):
			raise Exception("Cond.StartTimeAndDuration is missing! This is a mandatory condition.")
		else:
			cond = cls.StartTimeAndDuration
			parameters = RouteConditions[cond]
			LatestArrivalIn = parameters[3]
			if not CheckDurationWithNextConnection(ConnectionInfo, PathInfo, LatestArrivalIn):
				IncrementDicValue(cls.TerminationReasonsDic, 'StartTimeAndDuration_LatestArrival')
				if IfTest: print "--------- StartTimeAndDuration: LatestArrival (max duration) exceeded ---------"
				return None
			EarliestArrivalIn = parameters[2]
			if not CheckMinDurationWithNextConnection(ConnectionInfo, PathInfo, EarliestArrivalIn, EndStation, RouteConditions):
				IncrementDicValue(cls.TerminationReasonsDic, 'StartTimeAndDuration_EarliestArrival')
				if IfTest: print "--------- StartTimeAndDuration: EarliestArrival (min duration) violated ---------"
				return False

		# VisitStationsInGivenOrder
		if RouteConditions.has_key(cls.VisitStationsInGivenOrder):
			cond = cls.VisitStationsInGivenOrder
			parameters = RouteConditions[cond]
			OrderedStationList = parameters[0]
			if not CheckIfStationsAreVisitedInGivenOrder(ConnectionInfo, PathInfo, RouteConditions, OrderedStationList):
				IncrementDicValue(cls.TerminationReasonsDic, 'VisitStationsInGivenOrder')
				if IfTest: print "--------- VisitStationsInGivenOrder violated ---------"
				return False 

		# VisitAStationOnlyOnce
		if RouteConditions.has_key(cls.VisitAStationOnlyOnce):
			cond = cls.VisitAStationOnlyOnce
			IfVisitAStationOnlyOnce = RouteConditions[cond]
			if IfVisitAStationOnlyOnce and not CheckIfEachStationIsVisitedOnlyOnce(ConnectionInfo, PathInfo, EndStation, RouteConditions):
				IncrementDicValue(cls.TerminationReasonsDic, 'VisitAStationOnlyOnce')
				if IfTest: print "--------- VisitAStationOnlyOnce violated ---------"
				return False

		# VisitAConnectionOnlyOnce
		if RouteConditions.has_key(cls.VisitAConnectionOnlyOnce):
			cond = cls.VisitAConnectionOnlyOnce
			parameters = RouteConditions[cond]
			if CheckIfAConnectionIsVisitedOnlyOnce(ConnectionInfo, PathInfo, parameters, RouteConditions):
				IncrementDicValue(cls.TerminationReasonsDic, 'VisitAConnectionOnlyOnce')
				if IfTest: print "--------- VisitAConnectionOnlyOnce violated ---------"
				return False

		# SearchRoutesForEarliestArrival
		if cls.SearchRoutesForEarliestArrival in RouteConditions:
			cond = cls.SearchRoutesForEarliestArrival
			parameters = RouteConditions[cond]
			(CheckMinLineChange) = parameters
			if not CheckIfEarliestArrivalRouteSoFar(PathInfo, CheckMinLineChange):
				IncrementDicValue(cls.TerminationReasonsDic, 'SearchRoutesForEarliestArrival_RouteSelection')
				if IfTest: print "--------- SearchRoutesForEarliestArrival_RouteSelection violated ---------"
		
		# MaxNumberOfLineChanges
		if RouteConditions.has_key(cls.MaxNumberOfLineChanges):
			cond = cls.MaxNumberOfLineChanges
			parameters = RouteConditions[cond]
			MaxLineChangeLimit = parameters[0]
			if not CheckMaxNumberOfLineChanges(ConnectionInfo, PathInfo, MaxLineChangeLimit):
				IncrementDicValue(cls.TerminationReasonsDic, 'MaxNumberOfLineChanges')
				if IfTest: print "--------- MaxNumberOfLineChanges exceeded ---------"
				return False

		# ConnectionsAreAvailableOnAllListedDays
		if False and RouteConditions.has_key(cls.ConnectionsAreAvailableOnAllListedDays):
			cond = cls.ConnectionsAreAvailableOnAllListedDays
			parameters = RouteConditions[cond]
			DayList = parameters[0]
			if not CheckIfConnectionIsAvailableForAllListedDays(ConnectionInfo, DayList):
				IncrementDicValue(cls.TerminationReasonsDic, 'ConnectionsAreAvailableOnAllListedDays')
				if IfTest: print "--------- ConnectionsAreAvailableOnAllListedDays violated ---------"
				return False

		# MinStationCount
		if RouteConditions.has_key(cls.MinStationCount):
			cond = cls.MinStationCount
			parameters = RouteConditions[cond]
			MinStationCount = parameters[0]
			if not CheckMinStationCount(ConnectionInfo, PathInfo, MinStationCount, EndStation, RouteConditions):
				IncrementDicValue(cls.TerminationReasonsDic, 'MinStationCount')
				if IfTest: print "--------- MinStationCount violated ---------"
				return False

		# TimeForLineChange
		if not RouteConditions.has_key(cls.DefaultTimeForLineChange):
			raise Exception("Cond.TimeForLineChange is missing! This is a mandatory condition.")
		else:
			cond = cls.DefaultTimeForLineChange
			parameters = RouteConditions[cond]
			GeneralLineChangeTime = parameters[0]
			station = ConnectionInfo[ConnInfoInd['station_from']]

			# change for BU2019
			# MinChangeTime = GetLineChangeTimeAtStation(station, GeneralLineChangeTime)
			MinChangeTime = GeneralLineChangeTime

			if not CheckIfEnoughTimeForLineChange(ConnectionInfo, PathInfo, MinChangeTime):
				IncrementDicValue(cls.TerminationReasonsDic, 'TimeForLineChange')
				if IfTest: print "--------- TimeForLineChange violated ---------"
				return False

		# VisitStations
		if RouteConditions.has_key(cls.VisitStations):
			cond = cls.VisitStations
			parameters = RouteConditions[cond]
			StationList = parameters[0]
			IncludeOption = parameters[1]
			if IncludeOption in (INCLUDE_ALL_AND_ONLY, INCLUDE_ONLY):
				if not CheckIfStationIsInIncludeOnlyList(ConnectionInfo, StationList):
					IncrementDicValue(cls.TerminationReasonsDic, 'VisitStations')
					if IfTest: print "--------- VisitStations violated ---------"
					return False

		#ReturnFromCurrentStation
		if RouteConditions.has_key(cls.ReturnFromCurrentStation):
			cond=cls.ReturnFromCurrentStation
			IfReturnFromCurrentStation = RouteConditions[cond]
			if IfReturnFromCurrentStation and not CheckIfReturnIsPossibleFromCurrentStation(ConnectionInfo, PathInfo,RouteConditions,EarliestArrival_StationScoring):
				IncrementDicValue(cls.TerminationReasonsDic, 'ReturnFromCurrentStation')
				if IfTest: print "--------- ReturnFromCurrentStation violated ---------"
				return False

		# NoLineChangeAtVirtualStations like 138 (tunnel station)
		# don't permit line changes at virtual stations like 138
		if True:
			haltestelle_an = PathInfo[-1][ConnInfoInd['station_to']]
			if haltestelle_an < 1000:
				haltestelle_ab = ConnectionInfo[ConnInfoInd['station_from']]
				fahrt_id_last = PathInfo[-1][ConnInfoInd['travel_id']]
				fahrt_id_next = ConnectionInfo[ConnInfoInd['travel_id']]
				if haltestelle_ab < 1000 and fahrt_id_last != fahrt_id_next:
					return False


		# passed all conditions
		# Path is terminated successfuly if NextStation = EndStation
		if CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation):
			cls.RouteCountAfterConnectionSelection += 1
		return True

	@classmethod
	def CheckIfRouteShouldBeSelected(cls, PathInfo, RouteConditions):
		"""
		Determine whether the given route (PathInfo) should be selected 
		according to the conditions defined in dictionary RouteConditions.
		Returns:
		1: True (select --> include in the list of selected routes PathInfoList)
		2: False (deselect --> don't include in the list of selected routes PathInfoList)
		"""
		IfTest = cls.IfTestRouteSearch

		# VisitStations
		if RouteConditions.has_key(cls.VisitStations):
			cond = cls.VisitStations
			parameters = RouteConditions[cond]
			StationList = parameters[0]
			IncludeOption = parameters[1]
			if IncludeOption in (INCLUDE_ALL,):
				if not CheckIfRouteIncludesAllStationsInList(PathInfo, StationList):
					IncrementDicValue(cls.TerminationReasonsDic, 'VisitStations_INCLUDE_ALL')
					if IfTest: print "--------- VisitStations_INCLUDE_ALL violated ---------"
					return False

		# passed all conditions
		cls.RouteCountAfterRouteSelection += 1
		return True

	@classmethod
	def FilterRoutes(cls, PathInfoList, RouteConditions):
		"""
		Sort or select paths (routes) in PathInfoList
		"""
		IfTest = cls.IfTestRouteSearch

		Filtered_PathInfoList = PathInfoList

		# ShortestDurationPathSinceArrivalToFirstStation

		# SearchRoutesForEarliestArrival
		if cls.SearchRoutesForEarliestArrival in RouteConditions:
			cond = cls.SearchRoutesForEarliestArrival
			parameters = RouteConditions[cond]
			(CheckMinLineChange) = parameters
			Filtered_PathInfoList = SelectRoutesForEarliestArrival(Filtered_PathInfoList, CheckMinLineChange)

		# SortRoutesAfterValuesInDescOrder

		# final route count after filtering
		FinalRouteCount = len(Filtered_PathInfoList)
		cls.RouteCountAfterRouteFiltering = FinalRouteCount

		if IfTest or cls.ReportDuringRouteSearch in RouteConditions:
			print "%s routes remain after route filtering." % FinalRouteCount

		return Filtered_PathInfoList

# **************************************************************************************
# Core Dynamic Graph (Path Search) Algorithms
# **************************************************************************************

# global variables to store StationChain information
global_StationChainInfoPerFahrtID = {} 

def GetCompleteStationChainInformation(dbcur, RouteConditions):
	"""
	Read timetable data (Fahrplandaten) to set global variable global_StationChainInfoPerFahrtID
	that store station-chain information per line.

	global_StationChainInfoPerFahrtID[fahrt_id] = (
		[0] -> StationChain,			[st1, st2, ... stN]
		[1] -> DepartureTimeChain,		[t1, t2, ... None]
		[2] -> ArrivalTimeChain,		[None, t2, t3, ... tN]
		[3] -> LineIDChain,				(LineID of departure)
		[4] -> LineInfo 				(fahrt_id,fahrtnum,verwaltung,gattung,linie)
		)

	"""
	global global_StationChainInfoPerFahrtID 

	sql = """select fahrt_id,linie_id,fahrtnum,verwaltung,gattung,linie,
		haltestelle_ab,abfahrtm,haltestelle_an,ankunftm
		from %s where """ % tbl_TimeTable \
			+ Cond.GenerateSQLConditions(RouteConditions) +  " order by fahrt_id,hst_order;" 
	
	# execute sql
	dbcur.execute(sql)
	rows = dbcur.fetchall()

	PrevFahrtID = None
	PrevAnkunftm = None
	PrevStation = None

	for row in rows:
		fahrt_id = row[0]
		linie_id = row[1]
		fahrtnum = row[2]
		verwaltung = row[3]
		gattung = row[4]
		linie = row[5]
		haltestelle_ab = row[6]
		abfahrtm = row[7]
		haltestelle_an = row[8]
		ankunftm = row[9]

		if not fahrt_id in global_StationChainInfoPerFahrtID:
			# new fahrt_id
			LineInfo = (fahrt_id,fahrtnum,verwaltung,gattung,linie)
			
			# 0)StationChain, 1)DepartureTimeChain, 2)ArrivalTimeChain, 3)LineIDChain, 4)LineInfo
			global_StationChainInfoPerFahrtID[fahrt_id] = \
				([haltestelle_ab], [abfahrtm], [None], [linie_id], LineInfo)
			
			if PrevStation:
				# add to StationChainInfo of previous FahrtID
				global_StationChainInfoPerFahrtID[PrevFahrtID][0].append(PrevStation)
				global_StationChainInfoPerFahrtID[PrevFahrtID][1].append(None)
				global_StationChainInfoPerFahrtID[PrevFahrtID][2].append(PrevAnkunftm)
				global_StationChainInfoPerFahrtID[PrevFahrtID][3].append(None)
		else:
			# not new fahrt_id (i.e. not the first occurance of fahrt_id)
			global_StationChainInfoPerFahrtID[fahrt_id][0].append(haltestelle_ab)
			global_StationChainInfoPerFahrtID[fahrt_id][1].append(abfahrtm)
			global_StationChainInfoPerFahrtID[fahrt_id][2].append(PrevAnkunftm)
			global_StationChainInfoPerFahrtID[fahrt_id][3].append(linie_id)

		# assign previous values
		PrevFahrtID = fahrt_id 
		PrevAnkunftm = ankunftm 
		PrevStation = haltestelle_an

	# add to StationChainInfo of last FahrtID
	if PrevStation:
		global_StationChainInfoPerFahrtID[PrevFahrtID][0].append(PrevStation)
		global_StationChainInfoPerFahrtID[PrevFahrtID][1].append(None)
		global_StationChainInfoPerFahrtID[PrevFahrtID][2].append(PrevAnkunftm)
		global_StationChainInfoPerFahrtID[PrevFahrtID][3].append(None)

def ReadTimeTable(dbcur, RouteConditions):
	"""
	Read selected section of database table timetable into a list of N-tuples (ConnectionInfo),
	departure time index and Station-DepartureHour index.
	Returns: (TimeTableList, TimeTableIndex, StationHourIndex)
	"""
	# test: add conn_id for more deterministic ordering
	# 18.06.2017: add line_id <> '-1' condition for testing
	LineIDCond = " not (linie_id = '-1' and gattung='BUS' and linie='581') and "

	OrderedFields = ['conn_id','station_order','travel_id','travel_no','management','line_category','line','line_id','station_from','station_to',
		'departure_hour','departure_min','departure_totalmin','arrival_hour','arrival_min','arrival_totalmin','trafficdays_hexcode']
	OrderedFieldsStr = ','.join(OrderedFields)
	
	sql = """select %s from %s where """ % (OrderedFieldsStr, tbl_TimeTable) \
			+ Cond.GenerateSQLConditions(RouteConditions) +  " order by station_from,departure_totalmin,conn_id;" 
		 
	# execute sql
	dbcur.execute(sql)
	rows = dbcur.fetchall()
	RowCount = len(rows)

	# test
	# print "ReadTimeTable: SQL"
	# print sql 
	# print "RowCount: %s" % RowCount
	# quit()

	# init returned parameters
	TimeTableList = []
	TimeTableIndex = np.array(range(RowCount), int)
	StationHourIndex = {}

	# read all rows of table into TimeTable
	ind = 0
	CurStation = None
	CurHour = None

	for row in rows:
		# element order defined in ConnInfoInd
		l = range(0, len(ConnInfoInd))

		for i in range(0, len(OrderedFields)):
			FieldName = OrderedFields[i]
			FieldInd = OrderedFields.index(FieldName)
			l[ConnInfoInd[FieldName]] = row[FieldInd]
		t = tuple(l)
		TimeTableList.append(t)

		departure_hour = t[ConnInfoInd['departure_hour']]
		departure_min = t[ConnInfoInd['departure_min']]
		station_from = t[ConnInfoInd['station_from']]

		TimeTableIndex[ind] = 60*departure_hour + departure_min
		SHindex = (station_from, departure_hour)

		if not (CurStation == station_from and CurHour == departure_hour):
			StationHourIndex[SHindex] = ind
		CurStation = station_from
		CurHour = departure_hour
		ind += 1 
	return (TimeTableList, TimeTableIndex, StationHourIndex)


LMRequirementsAll = {
		('6.S10',2,11):   2,
		('4.S6b',2,11):   2,
		('7.S5a',2,11):   2,
		('7.S5b',2,11):   2,
		('4.S6a',2,11):   2,
		('8.S6',2,11):   2,
		('3.S6',2,11):   2,
		('3.RE2a',2,11):   2,
		('3.RE2b',2,11):   2,
		('3.S5',2,11):   2,
		('7.S14a',2,11):   2,
		('7.S14b',2,11):   2,
		('4.RE1',2,11):   2,
		('2.RE3',2,11):   2,
		('4.R5',2,11):   2,
		('4.R6',2,11):   2,
		('4.R7',2,11):   2,
		('2.S40',2,11):   2,
		('7.S3b',2,11):   2,
		('7.S3a',2,11):   2,
		('3.R1',2,11):   2,
		('3.RE4',2,11):   2,
		('3.R2',2,11):   2,
		('3.S2a',2,11):   2,
		('3.R7',2,11):   2,
		('3.R6',2,11):   2,
		('7.S25',2,11):   2,
		('3.S44c',2,11):   2,
		('7.S2a',2,11):   2,
		('7.S2b',2,11):   2,
		('1.R3',2,11):   2,
		('1.R4',2,11):   2,
		('3.S4b',2,11):   2,
		('5.RE1',2,11):   2,
		('5.S28',2,11):   2,
		('7.S16b',2,11):   2,
		('7.S16a',2,11):   2,
		('5.S9',2,11):   2,
		('3.S51',2,11):   2,
		('3.S52',2,11):   2,
		('6.S30',2,11):   2,
		('2.RE4a',2,11):   2,
		('2.RE4b',2,11):   2,
		('6.S3',2,11):   2,
		('6.S2',2,11):   2,
		('6.S9',2,11):   2,
		('2.S1',2,11):   2,
		('7.S8c',2,11):   2,
		('7.S8b',2,11):   2,
		('7.S8a',2,11):   2,
		('6.S20',2,11):   2,
		('2.S30',2,11):   2,
		('6.S25',2,11):   2,
		('6.S26',2,11):   2,
		('6.S27',2,11):   2,
		('6.S29',2,11):   2,
		('1.S2b',2,11):   2,
		('1.S2a',2,11):   2,
		('6.S23a',2,11):   2,
		('6.S23b',2,11):   2,
		('6.S41',2,11):   2,
		('2.R8',2,11):   2,
		('3.S3b',2,11):   2,
		('3.S3a',2,11):   2,
		('3.S4a',2,11):   2,
		('5.S1',2,11):   2,
		('2.R11',2,11):   2,
		('3.RE1a',2,11):   2,
		('5.S8',2,11):   2,
		('5.S3a',2,11):   2,
		('5.S3b',2,11):   2,
		('5.S3c',2,11):   2,
		('6.RE2',2,11):   2,
		('7.S7b',2,11):   2,
		('7.S7a',2,11):   2,
		('3.S2b',2,11):   2,
		('2.S9',2,11):   2,
		('3.R4b',2,11):   2,
		('3.R4a',2,11):   2,
		('7.S15b',2,11):   2,
		('7.S15a',2,11):   2,
		('2.S4b',2,11):   2,
		('2.S4a',2,11):   2,
		('7.S24b',2,11):   2,
		('7.S24c',2,11):   2,
		('7.S24a',2,11):   2,
		('2.R12',2,11):   2,
		('7.S6a',2,11):   2,
		('3.S44a',2,11):   2,
		('3.S44b',2,11):   2,
		('7.S6b',2,11):   2,
		('3.S1a',2,11):   2,
		('3.S1b',2,11):   2,
		('3.RE1b',2,11):   2,
		('2.R10',2,11):   2,
		('7.S9a',2,11):   2,
		('7.S9b',2,11):   2,
		('2.R13',2,11):   2,
		('2.R14',2,11):   2,
		('7.S12b',2,11):   2,
		('7.S12a',2,11):   2,
		('1.S3a',2,11):   2,
		('1.S3b',2,11):   2,
		('8.RE1',2,11):   2,
		('6.S10',3,11):   2,
		('4.S6b',3,11):   2,
		('7.S5a',3,11):   2,
		('7.S5b',3,11):   2,
		('4.S6a',3,11):   2,
		('8.S6',3,11):   2,
		('3.S6',3,11):   2,
		('3.RE2a',3,11):   2,
		('3.RE2b',3,11):   2,
		('3.S5',3,11):   2,
		('7.S14a',3,11):   2,
		('7.S14b',3,11):   2,
		('4.RE1',3,11):   2,
		('2.RE3',3,11):   2,
		('4.R5',3,11):   2,
		('4.R6',3,11):   2,
		('4.R7',3,11):   2,
		('2.S40',3,11):   2,
		('7.S3b',3,11):   2,
		('7.S3a',3,11):   2,
		('3.R1',3,11):   2,
		('3.RE4',3,11):   2,
		('3.R2',3,11):   2,
		('3.S2a',3,11):   2,
		('3.R7',3,11):   2,
		('3.R6',3,11):   2,
		('7.S25',3,11):   2,
		('3.S44c',3,11):   2,
		('7.S2a',3,11):   2,
		('7.S2b',3,11):   2,
		('1.R3',3,11):   2,
		('1.R4',3,11):   2,
		('3.S4b',3,11):   2,
		('5.RE1',3,11):   2,
		('5.S28',3,11):   2,
		('7.S16b',3,11):   2,
		('7.S16a',3,11):   2,
		('5.S9',3,11):   2,
		('3.S51',3,11):   2,
		('3.S52',3,11):   2,
		('6.S30',3,11):   2,
		('2.RE4a',3,11):   2,
		('2.RE4b',3,11):   2,
		('6.S3',3,11):   2,
		('6.S2',3,11):   2,
		('6.S9',3,11):   2,
		('2.S1',3,11):   2,
		('7.S8c',3,11):   2,
		('7.S8b',3,11):   2,
		('7.S8a',3,11):   2,
		('6.S20',3,11):   2,
		('2.S30',3,11):   2,
		('6.S25',3,11):   2,
		('6.S26',3,11):   2,
		('6.S27',3,11):   2,
		('6.S29',3,11):   2,
		('1.S2b',3,11):   2,
		('1.S2a',3,11):   2,
		('6.S23a',3,11):   2,
		('6.S23b',3,11):   2,
		('6.S41',3,11):   2,
		('2.R8',3,11):   2,
		('3.S3b',3,11):   2,
		('3.S3a',3,11):   2,
		('3.S4a',3,11):   2,
		('5.S1',3,11):   2,
		('2.R11',3,11):   2,
		('3.RE1a',3,11):   2,
		('5.S8',3,11):   2,
		('5.S3a',3,11):   2,
		('5.S3b',3,11):   2,
		('5.S3c',3,11):   2,
		('6.RE2',3,11):   2,
		('7.S7b',3,11):   2,
		('7.S7a',3,11):   2,
		('3.S2b',3,11):   2,
		('2.S9',3,11):   2,
		('3.R4b',3,11):   2,
		('3.R4a',3,11):   2,
		('7.S15b',3,11):   2,
		('7.S15a',3,11):   2,
		('2.S4b',3,11):   2,
		('2.S4a',3,11):   2,
		('7.S24b',3,11):   2,
		('7.S24c',3,11):   2,
		('7.S24a',3,11):   2,
		('2.R12',3,11):   2,
		('7.S6a',3,11):   2,
		('3.S44a',3,11):   2,
		('3.S44b',3,11):   2,
		('7.S6b',3,11):   2,
		('3.S1a',3,11):   2,
		('3.S1b',3,11):   2,
		('3.RE1b',3,11):   2,
		('2.R10',3,11):   2,
		('7.S9a',3,11):   2,
		('7.S9b',3,11):   2,
		('2.R13',3,11):   2,
		('2.R14',3,11):   2,
		('7.S12b',3,11):   2,
		('7.S12a',3,11):   2,
		('1.S3a',3,11):   2,
		('1.S3b',3,11):   2,
		('8.RE1',3,11):   2,
}


def FindAllRoutes(dbcur, RouteConditions,Requirements,EarliestArrival,ClusterIndex):
	"""
	Find all possible routes (w.r.t. time table) from start to end station 
    according to all conditions given in RouteConditions. 
    
    RouteConditions: Dictionary containing all route conditions
    including start and end stations.
    
    Return: PathInfoList (synonym for RouteInfoList)
	"""

	if not RouteConditions:
		return None

	# get path start
	# StartTimeAndDuration: (9, 45, 1*60, 6*60)
	PathBeginTimeHour = RouteConditions[Cond.StartTimeAndDuration][0]
	PathBeginTimeMin = RouteConditions[Cond.StartTimeAndDuration][1]
	StartStation = RouteConditions[Cond.StartAndEndStations][0]
	EndStation = RouteConditions[Cond.StartAndEndStations][1]

	if Cond.TestRunDuringRouteSearch in RouteConditions:
		Cond.IfTestRouteSearch = True 
		Cond.TestWaitingTime = RouteConditions[Cond.TestRunDuringRouteSearch][0]

	# create first ConnectionInfo of path
	# NOTE: departure time = arrival time, path duration is calculated w.r.t. first departure time!!!
	ConnectionInfo = range(0, len(ConnInfoInd))		# note: alias keys in ConnInfoInd
	for key in ConnInfoInd:
		if key == 'station_from':
			ConnectionInfo[ConnInfoInd['station_from']] = 8500000
		elif key == 'station_to':
			ConnectionInfo[ConnInfoInd['station_to']] = StartStation
		elif key == 'line_category':
			ConnectionInfo[ConnInfoInd['line_category']] = 'W'
		elif key == 'departure_hour':
			ConnectionInfo[ConnInfoInd['departure_hour']] = PathBeginTimeHour
		elif key == 'departure_min':
			ConnectionInfo[ConnInfoInd['departure_min']] = PathBeginTimeMin	
		elif key == 'arrival_hour':
			ConnectionInfo[ConnInfoInd['arrival_hour']] = PathBeginTimeHour
		elif key == 'arrival_min':
			ConnectionInfo[ConnInfoInd['arrival_min']] = PathBeginTimeMin	
		else:
			ConnectionInfo[ConnInfoInd[key]] = None
	ConnectionInfo = tuple(ConnectionInfo)

	# read table with RouteConditions
	print "START reading timetable data (Fahrplandaten) from database..."
	st = time.time()
	(TimeTableList, TimeTableIndex, StationHourIndex) = ReadTimeTable(dbcur, RouteConditions)
	print "FINISHED reading timetable data (Fahrplandaten) from database, in %.2f seconds." % (time.time() - st)
	print "TEST: SizeOf variable TimeTableList in kilobytes: %d" % math.floor(sys.getsizeof(TimeTableList) / 2**10)

	if Cond.IfTestRouteSearch:
		print "\nTimeTableList with haltestelle_ab (Start Station) = %s" % StartStation
		for c in TimeTableList: 
			if c[0] == StartStation:
				print c
		N = 20
		print "\nLast %s entries of TimeTableList:" % N
		L = len(TimeTableList)
		for i in range(L-N,L):
			print TimeTableList[i]

	#DistinctLineID and DistinctStation Set Generation
	arr=np.array(TimeTableList)
	DistinctStations=np.unique(arr[:,ConnInfoInd['station_from']])
	b=np.unique(arr[:,ConnInfoInd['station_to']])
	DistinctStations=set(DistinctStations)
	b=set(b)
	DistinctStations.update(b)

	DistinctLineID=np.unique(arr[:,ConnInfoInd['line_id']])
	DistinctLineID = [x for x in DistinctLineID if x is not None and x is not '-1']
	StationListForLines=dict() #Dictionary Containing Stations that Lines Operate

	for line in DistinctLineID:

		StationList=set()

		for connection in arr:

			if connection[ConnInfoInd['line_id']]==line:
				StationList.add(connection[ConnInfoInd['station_from']])
				StationList.add(connection[ConnInfoInd['station_to']])

		StationListForLines[line] = StationList
	
	# workbook = openpyxl.Workbook()
	# sheet = workbook.active

	# # openpyxl does things based on 1 instead of 0
	# row = 1
	# for key,values in StationListForLines.items():
	# 	# Put the key in the first column for each key in the dictionary
	# 	sheet.cell(row=row, column=1, value=key)
	# 	column = 2
	# 	for element in values:
	# 		# Put the element in each adjacent column for each element in the tuple
	# 		sheet.cell(row=row, column=column, value=element)
	# 		column += 1
	# 	row += 1

	# workbook.save(filename="StationListForLines.xlsx")

	#Creating EarliestArrival_StationScoring data frame to be used as a criteria for turning back
	EarliestArrival_StationScoring=pd.DataFrame(np.full((1, len(DistinctStations)), 0),columns=list(DistinctStations))

	for station in DistinctStations:
		EarliestArrival_StationScoring.at[0,station]=EarliestArrival.at[station,StartStation]

	# conn_id=[]
	# for rows in TimeTableList:
	# 	conn_id.append(rows[ConnInfoInd['conn_id']])

	# RequirementsSet=set(list(list(zip(*LMRequirementsAll)[0])))
	# RequirementsSet=list(RequirementsSet)

	#Initializing connection scoring dataframe with respect to requirements
	# global RequirementScores
	# RequirementScores = pd.DataFrame(np.full((len(TimeTableList),len(RequirementsSet)),None),columns=RequirementsSet)
	# RequirementScores['conn_id']=conn_id
	# RequirementScores=RequirementScores.set_index('conn_id',drop=True)

	# DistanceMatrix=pd.read_excel("DistanceMatrix.xlsx",index_col=0)
	#Initializing EarliestArrival to be used in connection scoring
	
	#EarliestArrival=pd.DataFrame(np.full((len(DistinctStations), len(DistinctStations)), 1000),columns=list(DistinctStations))
	#EarliestArrival['StationFrom']=list(DistinctStations)
	#EarliestArrival=EarliestArrival.set_index('StationFrom',drop=True)
	

	# def StationScoring(StationList,ScoredStations,StartingStation):
	# 	#StationList: Stations which we are scoring(starting station is given at the start)
		
	# 	if len(StationList) is 0:
	# 		return
		
	# 	global EarliestArrival

	# 	Start = next(iter(StartingStation))

	# 	if StationList == StartingStation:
	# 		EarliestArrival.at[Start,Start]=0
		
	# 	if ScoredStations == DistinctStations:
	# 		return #Returns final EarliestArrival for specific station
		
	# 	if len(ScoredStations) is 0:
	# 		ScoredStations=ScoredStations.union(StationList) 
	# 	else:
	# 		ScoredStations.update(StationList)
		
	# 	Intersection=set()

	# 	for station in StationList:

	# 		for connection_row in TimeTableList:
				
	# 			if connection_row[ConnInfoInd['station_from']]==station:

	# 				if connection_row[ConnInfoInd['station_to']] not in ScoredStations:
	# 					Intersection.add(connection_row[ConnInfoInd['station_to']])

	# 				if EarliestArrival.at[Start,connection_row[ConnInfoInd['station_to']]] > EarliestArrival.at[Start,connection_row[ConnInfoInd['station_from']]] + connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]:
	# 					EarliestArrival.at[Start,connection_row[ConnInfoInd['station_to']]] = EarliestArrival.at[Start,connection_row[ConnInfoInd['station_from']]] + connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]

	# 	StationScoring(Intersection,ScoredStations,StartingStation)

	# def ConnectionScoring(StationList,CoveredStations):
	# 	#StationList: Stations of lines which we want to cover (trial is used as an example)
	# 	#level= neighborhood level of all other connections
	# 	# CoveredStations = Set of so far Covered stations
	# 	global RequirementScores
	# 	global EarliestArrival

	# 	if CoveredStations == DistinctStations:
	# 		return #Returns final(with final connection scores) TimeTableList

	# 	if len(CoveredStations) is 0:
	# 		CoveredStations=CoveredStations.union(StationList) 
	# 	else:
	# 		CoveredStations.update(StationList)
		
	# 	Intersection=set()

	# 	for station in StationList:

	# 		for connection_row in TimeTableList:
				
	# 			if connection_row[ConnInfoInd['station_from']]==station:

	# 				if connection_row[ConnInfoInd['station_to']] not in CoveredStations:
	# 					Intersection.add(connection_row[ConnInfoInd['station_to']])

	# 				if (
	# 					RequirementScores.at[connection_row[ConnInfoInd['conn_id']],requirement]==None  or 
	# 					RequirementScores.at[connection_row[ConnInfoInd['conn_id']],requirement]> EarliestArrival.at[0,connection_row[ConnInfoInd['station_from']]]+connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]
	# 					):
	# 					RequirementScores.at[connection_row[ConnInfoInd['conn_id']],requirement]=EarliestArrival.at[0,connection_row[ConnInfoInd['station_from']]]+connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]
					
	# 				if EarliestArrival.at[0,connection_row[ConnInfoInd['station_to']]] > EarliestArrival.at[0,connection_row[ConnInfoInd['station_from']]] + connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]:
	# 					EarliestArrival.at[0,connection_row[ConnInfoInd['station_to']]] = EarliestArrival.at[0,connection_row[ConnInfoInd['station_from']]] + connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]

	# 	ConnectionScoring(Intersection,CoveredStations)
	
	#start_time1 = timeit.default_timer()
	#Filling out the EarliestArrival
	# for station in DistinctStations:
	# 	start_time2 = timeit.default_timer()
	# 	ScoredStations=set()
	# 	StationScoring({station},ScoredStations,{station})
	# 	elapsed2 = timeit.default_timer() - start_time2
	# 	print 'station %d earliest arrival takes %f ' %(station,elapsed2)
	# elapsed1 = timeit.default_timer() - start_time1
	# print 'filling out earliest arrival takes %f ' %(elapsed1)
	#EarliestArrival.to_excel("earliest_arrival_allstations.xlsx")
	
	# start_time1 = timeit.default_timer()
	# #Filling out the Requirements Score dataframe 
	# for requirement in RequirementsSet:
	# 	start_time2 = timeit.default_timer()
	# 	Stations=StationListForLines[requirement]
	# 	Stations=list(Stations)
	
	# 	#Setting the earliest arrival to "0" for requirement line's stations
	# 	#EarliestArrival.loc[:,Stations]=0

	# 	#Setting the connection score to "0" for requirement line's connections
	# 	for connection_row in TimeTableList:
	# 		if connection_row[ConnInfoInd['line_id']] == requirement:
	# 			RequirementScores.at[connection_row[ConnInfoInd['conn_id']],requirement]=0
	# 		else:
	# 			RequirementScores.at[connection_row[ConnInfoInd['conn_id']],requirement]=DistanceMatrix.loc[connection_row[ConnInfoInd['station_to']],Stations].mean()+connection_row[ConnInfoInd['arrival_totalmin']]-connection_row[ConnInfoInd['departure_totalmin']]
	# 	elapsed2 = timeit.default_timer() - start_time2
	# 	print 'line %s reqscore takes %f ' %(requirement,elapsed2)
	
	# elapsed1 = timeit.default_timer() - start_time1
	# print 'initial reqscore takes %f ' %(elapsed1)

	# RequirementScores.to_excel("reqscores_with_coordinates.xlsx")

	#Filling out the EarliestArrival_StationScoring

	# for station in DistinctStations:
	# 	start_time1 = timeit.default_timer()
	# 	global EarliestArrivalSpecificStation
	# 	EarliestArrivalSpecificStation=pd.DataFrame(np.full((1, len(DistinctStations)), 1e8),columns=list(DistinctStations))
	# 	ScoredStations=set()
	# 	StationScoring({station},ScoredStations,{station})
	# 	EarliestArrival_StationScoring.at[0,station]=EarliestArrivalSpecificStation.at[0,StartStation]
	# 	elapsed1 = timeit.default_timer() - start_time1
	# 	print 'Station %f earliest arrival takes %f ' %(station,elapsed1)
	
	
	# EarliestArrival_StationScoring.to_excel("earliest_arrival.xlsx")
	# RequirementScores_numpy=RequirementScores.to_numpy(copy=True)
	# RequirementScores_numpy[np.isnan(RequirementScores_numpy)] = 0
	# wcss = []
	# for i in range(1, 20):
	# 	kmeans = KMeans(n_clusters=i, init='k-means++', max_iter=300, n_init=10, random_state=0)
	# 	#print np.any(np.isnan(RequirementScores_numpy))
	# 	#print np.all(np.isfinite(RequirementScores_numpy))
	# 	kmeans.fit(np.transpose(RequirementScores_numpy))
	# 	wcss.append(kmeans.inertia_)

	# plt.plot(range(1, 20), wcss)
	# plt.title('Elbow Method')
	# plt.xlabel('Number of clusters')
	# plt.ylabel('WCSS')
	# plt.show()
	
	# ChosenNumberOfClusters=4

	# kmeans=KMeans(n_clusters=ChosenNumberOfClusters, init='k-means++', max_iter=300, n_init=10, random_state=0)
	# kmeans.fit(np.transpose(RequirementScores_numpy))
	# workbook = openpyxl.Workbook()
	# sheet = workbook.active
	
	# for i in range(1,len(kmeans.labels_)):

	# 	sheet.cell(row=i, column=1, value=RequirementScores.columns.values[i-1])
	# 	sheet.cell(row=i, column=2, value=kmeans.labels_[i-1])
	
	# workbook.save(filename="clusters.xlsx")

	RequirementsSet=set(list(list(zip(*Requirements)[0])))
	RequirementsSet=list(RequirementsSet)

	# find all possible paths
	FindAllRoutesRec(ConnectionInfo, EndStation, RouteConditions, \
		TimeTableList, TimeTableIndex,StationHourIndex,RequirementsSet,EarliestArrival_StationScoring,ClusterIndex)
	
	PathInfoList = Cond.SelectedRoutes

	# apply filter
	RouteInfoList = Cond.FilterRoutes(PathInfoList, RouteConditions)

	# reset class variables
	(StatusReport, TerminationReasons) = Cond.ResetClassVariables() 
	return (RouteInfoList, StatusReport, TerminationReasons)

def FindAllRoutesRec(ConnectionInfo, EndStation, RouteConditions, TimeTableList, TimeTableIndex, StationHourIndex,RequirementsList,EarliestArrival_StationScoring,ClusterIndex,PathInfo=[]):
	""" 
	Find all possible routes (w.r.t. time table) from start to end station w.r.t.
	all conditions given by the dictionary RouteConditions.
	"""

	PathInfo = PathInfo + [ConnectionInfo]

	if Cond.IfTestRouteSearch:
		Stations = GetAllStationsOfRoute(PathInfo)
		print "\nStations of Path (%s): ++++++++" % len(Stations)
		print Stations
		print "Route Information:"
		print PrettyStringRouteInfo(PathInfo)

	# check successful termination and measurement
    # if len(PathInfo) > 1 and ConnectionInfo[ConnInfoInd['station_to']] == EndStation:  
	
	if CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation):
		#start_time = timeit.default_timer()
		if Cond.IfTestRouteSearch:
			print "End Station is reached!"	
		
		(_, LMCoveragePerLineKey) = \
			GetLMCoverageOfRoute(PathInfo, 10, PeriodBegin, PeriodEnd, LMRequirements=LMRequirementsAll)
		
		if len(LMCoveragePerLineKey)!=0:

			measured_lines=set(list(list(zip(*LMCoveragePerLineKey)[0])))
			measured_lines=list(measured_lines)

			for lines in measured_lines:
				if lines in list(RequirementScoresWithinCluster_dict[ClusterIndex].columns.values):
					RequirementScoresWithinCluster_dict[ClusterIndex].loc[:,lines]=None
					#print "Line %s  is measured" %(lines)
		#elapsed = timeit.default_timer() - start_time
		#print 'whole measurement check takes %f seconds' %(elapsed)
		return [PathInfo]
	
    # current (this iteration's) path length
	CurPathLen = len(PathInfo)

    # get next connections
	start_station = ConnectionInfo[ConnInfoInd['station_to']]
	departure_hour = ConnectionInfo[ConnInfoInd['arrival_hour']] 	
	departure_min = ConnectionInfo[ConnInfoInd['arrival_min']]

    # TEST BU2019
	if False:
		print 'ConnInfoInd: ' + str(ConnectionInfo)
		print 'start_station,departure_hour,departure_min: %s, %s, %s' % (start_station, departure_hour, departure_min)
		time.sleep(0.1)
    
    # mandatory conditions
	WaitLimit = RouteConditions[Cond.MaxWaitingTimeAtStation][0]
    
    # get next connections from the station
	ConnectionInfoList = GetListOfNextConnections(TimeTableList, TimeTableIndex, StationHourIndex, start_station, departure_hour, departure_min, WaitLimit)

	if Cond.IfTestRouteSearch:
		print "Next connections:"
		for c in ConnectionInfoList:
			print c
		time.sleep(Cond.TestWaitingTime)

	if not ConnectionInfoList:		# Endstation: Node w/o successor nodes
		return []

	PathInfoList = []
	res=[]
	
	for ConnectionInfo in ConnectionInfoList:
		res.append(Cond.CheckIfConnectionShouldBeSelected(ConnectionInfo, PathInfo, EndStation, RouteConditions,EarliestArrival_StationScoring))
		
		if res[-1] == None: return[]

		if res[-1] == True:
			connection_id=ConnectionInfo[ConnInfoInd['conn_id']]
			res[-1]= RequirementScoresWithinCluster_dict[ClusterIndex][RequirementScoresWithinCluster_dict[ClusterIndex]!=None].min(1)[connection_id]
		
		else:

			res[-1]=-1

	b=ConnectionInfoList

	c= [i for i, n in enumerate(res) if n == -1]
	for index in sorted(c, reverse=True):
		del b[index]

	OrderedConnections=[b for _,b in sorted(zip(res,b),reverse=False)]
	#PriorityConnection=ConnectionInfoList[res.index(max(res))]

	for ConnectionInfo in OrderedConnections:
		
		# recursive call
		extended_paths = FindAllRoutesRec(ConnectionInfo, EndStation, RouteConditions, \
			TimeTableList, TimeTableIndex, StationHourIndex,RequirementsList, EarliestArrival_StationScoring,ClusterIndex,PathInfo)

		# report status
		if Cond.ReportDuringRouteSearch in RouteConditions:
			TimeIntv = default_timer() - Cond.SearchStartTime
			RouteSearchReportingIntervalInSeconds = RouteConditions[Cond.ReportDuringRouteSearch][0]
			if TimeIntv > Cond.RouteSearchReportCounter * RouteSearchReportingIntervalInSeconds:
				Cond.RouteSearchReportCounter += 1 
				print "%s seconds passed... " % "{:.2f}".format(TimeIntv)
				print "%s routes found so far, that passed all connection selection criteria (before route selection)" \
					% Cond.RouteCountAfterConnectionSelection	
				print "%s routes found so far, that passed all route selection criteria (before final route filtering)" \
					% Cond.RouteCountAfterRouteSelection	
				print "----------------------"	

		# append to path list
		for p in extended_paths:
			# no need to recheck route unless current connection is the last one 
			# LastConnection = (ConnectionInfo == p[-1])
			LastConnection = (CurPathLen == len(p) -1 and ConnectionInfo == p[-1])
				
			if LastConnection:

				if Cond.CheckIfRouteShouldBeSelected(p, RouteConditions):
					PathInfoList.append(p)
					Cond.SelectedRoutes.append(ApplyAllRouteInfoCorrections(p))

					# evaluate route
					# cancel for BU2019

					if Cond.IfTestRouteSearch:
						print "%s routes found so far, that passed all connection selection criteria (before route selection)" \
							% Cond.RouteCountAfterConnectionSelection
						print "%s routes found so far, that passed all route selection criteria (before final route filtering)\n" \
							% Cond.RouteCountAfterRouteSelection		
						print "----------------------"	

					# test
					IncrementDicValue(Cond.RouteCountPerRouteLength, CurPathLen)
						
				else:
					# not last connection, no need to recheck the route
					# PathInfoList.append(p)
					# IncrementDicValue(SelectedRoutesPerLevel, CurPathLen)
					pass
		
	return PathInfoList

# **************************************************************************************
# Path/Connection Evaluation Functions
# **************************************************************************************

def GetDurationUntilDeparture(NextConnectionInfo, PathInfo):
	"""
	Return total duration of (path + time until departure for next station) in minutes.
	Duration since arrival to first station (Depot).
	"""
	departure_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	departure_next_station = NextConnectionInfo[ConnInfoInd['departure_hour']]*60 + NextConnectionInfo[ConnInfoInd['departure_hour']]
	return (departure_next_station - departure_first_station)

def GetDurationUntilNextArrival(NextConnectionInfo, PathInfo):
	"""
	Return total duration of (path + time until arrival at next station) in minutes.
	Duration since arrival to first station (Depot).
	"""
	departure_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	arrival_next_station = NextConnectionInfo[ConnInfoInd['arrival_hour']]*60 + NextConnectionInfo[ConnInfoInd['arrival_hour']]
	return (arrival_next_station - departure_first_station)

def GetTotalDurationOfPath(PathInfo):
	"""
	Get total duration of path in seconds since arrival to first station (Depot).
	"""
	if PathInfo == None: return 0
	if len(PathInfo) < 2: return 0
	
	departure_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	# departure_first_station = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]

	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	return (arrival_last_station - departure_first_station)

GetTotalDurationOfPathSinceSearchStart = GetTotalDurationOfPath

def GetTimeAndDurationOfPath(PathInfo):
	"""
	Get begin/end times, and total duration of path since arrival to first station (Depot).
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None 

	arrival_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	# departure_first_station = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]

	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	TotalDuration = arrival_last_station - arrival_first_station
	return (TotalDuration, arrival_first_station, arrival_last_station)

GetTimeAndDurationOfPathSinceSearchStart = GetTimeAndDurationOfPath

def GetDurationOfTripSinceDeparture(PathInfo):
	"""
	Get total duration of trip (in minutes) since departure from the first station (Depot).
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None 

	departure_first_station = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]

	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	TotalDuration = arrival_last_station - departure_first_station
	return TotalDuration

# GetDurationOfTour
GetDurationOfTour = GetDurationOfTripSinceDeparture 

def GetDepartureTimeOfTour(PathInfo):
	"""
	Get departure time of tour (departure from the first station) in total minutes.
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None  

	DepartureTime = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]
	return DepartureTime 

def GetArrivalTimeOfTour(PathInfo):
	"""
	Get arrival time of tour (arrival to last station) in total minutes.
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None  

	ArrivalTime = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	return ArrivalTime 

def GetTimeAndDurationOfTripSinceDeparture(PathInfo):
	"""
	Get begin/end times, and total duration of path since departure from the first station (Depot).
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None 

	departure_first_station = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]

	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	TotalDuration = arrival_last_station - departure_first_station
	return (TotalDuration, departure_first_station, arrival_last_station)

def GetListOfNextConnections(TimeTableList, TimeTableIndex, StationHourIndex, CurrentStation, 
	abfahrt_std, abfahrt_min, WaitingTime):
	"""
	Get list of next stations (list of ConnectionInfo) from time TimeTableList 
	within the given time interval (WaitingTime in minutes).
	"""
	interval_start = 60*abfahrt_std + abfahrt_min
	interval_end = interval_start + WaitingTime

	FilteredTimeTable = []
	hour = abfahrt_std

	# what if key does not exist (i.e. no connection in data set)
	if not StationHourIndex.has_key((CurrentStation, abfahrt_std)):
		# f.e. no connection found after 7.55 until 8:00
		
		if StationHourIndex.has_key((CurrentStation, abfahrt_std+1)):
			hour += 1
		else:
			return FilteredTimeTable

	for i in range(StationHourIndex[(CurrentStation, hour)], len(TimeTableList)):
		if TimeTableIndex[i] <= interval_end and TimeTableList[i][ConnInfoInd['station_from']] == CurrentStation:
			if TimeTableIndex[i] >= interval_start:
				FilteredTimeTable.append(TimeTableList[i])
		else:
			break
	return FilteredTimeTable

def GetAllStationsOfRoute(PathInfo):
	"""
	Get route as time-ordered list of stations from PathInfo
	"""
	path = []
	for ConnectionInfo in PathInfo:
		path.append(ConnectionInfo[ConnInfoInd['station_from']])
	if len(PathInfo) > 0:
		path.append(PathInfo[-1][ConnInfoInd['station_to']])
	return path

def GetTimeIntervalForMeasurementOnLine(Line, TimeIntervalOnLine, MeasureIntervalBegin, MeasureIntervalEnd, RequiredMeasureTime):
	"""
	Return time interval for measurement if given line can be measured
	within the given time interval (TimeIntervalBegin, TimeIntervalEnd); otherwise None.
	
	TimeIntervalOnLine: Time intervals spent on each line (route information)
	RequiredMeasureTime: Time in minutes required for a measurement
	None for TimeIntervalBegin/TimeIntervalEnd  means there is no lower/upper limit for time interval.

	NOTE:
	First time interval is chosen in case there are multiple time intervals 
	on the line that would be sufficiently long for a measurement.
	"""
	if not TimeIntervalOnLine:
		return None

	if not TimeIntervalOnLine.has_key(Line):
		return None

	TimeIntervals = TimeIntervalOnLine[Line]

	# check lower and upper time limits
	if MeasureIntervalBegin == None: MeasureIntervalBegin = 0 
	if MeasureIntervalEnd == None: MeasureIntervalEnd = 60*60

	for TimeInt in TimeIntervals:
		TimeIntBegin = TimeInt[0]
		TimeIntEnd = TimeInt[1]

		TimeIntersection = GetIntersectionLengthOfTwoLines(TimeIntBegin, TimeIntEnd, MeasureIntervalBegin, MeasureIntervalEnd)
		
		if TimeIntersection >= RequiredMeasureTime:
			MTimeIntvBegin = max(TimeIntBegin, MeasureIntervalBegin)
			MTimeIntvEnd = MTimeIntvBegin + TimeIntersection
			TimeIntvForMeasurement = (MTimeIntvBegin, MTimeIntvEnd)
			return TimeIntvForMeasurement
	
	return None

def GetTimeIntervalForMeasurementAtStation(Station, TimeIntervalAtStation, MeasureIntervalBegin, MeasureIntervalEnd, RequiredMeasureTime):
	"""
	Return time interval for measurement if given station can be measured
	within the given time interval (TimeIntervalBegin, TimeIntervalEnd); otherwise None.
	
	TimeIntervalAtStation: Time intervals spent at each station (route information)
	RequiredMeasureTime: Time in minutes required for a measurement
	None for TimeIntervalBegin/TimeIntervalEnd  means there is no lower/upper limit for time interval.

	NOTE:
	First time interval is chosen in case there are multiple time intervals 
	at station that would be sufficiently long for a measurement.
	"""
	if not TimeIntervalAtStation:
		return None

	if not TimeIntervalAtStation.has_key(Station):
		return None

	TimeIntervals = TimeIntervalAtStation[Station]

	# check lower and upper time limits
	if MeasureIntervalBegin == None: MeasureIntervalBegin = 0 
	if MeasureIntervalEnd == None: MeasureIntervalEnd = 60*60

	for TimeInt in TimeIntervals:
		TimeIntBegin = TimeInt[0]
		TimeIntEnd = TimeInt[1]

		TimeIntersection = GetIntersectionLengthOfTwoLines(TimeIntBegin, TimeIntEnd, MeasureIntervalBegin, MeasureIntervalEnd)
		
		if TimeIntersection >= RequiredMeasureTime:
			MTimeIntvBegin = max(TimeIntBegin, MeasureIntervalBegin)
			MTimeIntvEnd = MTimeIntvBegin + TimeIntersection
			TimeIntvForMeasurement = (MTimeIntvBegin, MTimeIntvEnd)
			return TimeIntvForMeasurement
	
	return None

def GetAllLinesOfPath(PathInfo):
	"""
	Return all lines of path as a set of line IDs
	"""
	LineSet = set()
	for ConnectionInfo in PathInfo:
		line_id = ConnectionInfo[ConnInfoInd['line_id']]
		LineSet.add(line_id)
	return LineSet

def GetAllLinesOfPathWithinTimeWindow(PathInfo, TWindow):
	"""
	Return all lines of path (whose departure times 
	lie within the given time window) as a set of line IDs.
	
	TWindow is an interval like (5*60, 8*60)
	"""
	LineSet = set()
	for ConnectionInfo in PathInfo:
		LineID = ConnectionInfo[ConnInfoInd['line_id']]
		DepartureTime = 60* ConnectionInfo[ConnInfoInd['departure_hour']] + ConnectionInfo[ConnInfoInd['departure_min']]
		
		if DepartureTime >= TWindow[0] and DepartureTime <= TWindow[1]:
			LineSet.add(LineID)
	
	return LineSet

def GetOrderedLinesOfPath(PathInfo):
	"""
	Return all lines of path as a time-ordered list of line IDs.
	Adds a single line for a series of connections on the same line.
	"""
	LineList = []

	for ConnectionInfo in PathInfo:
		line_id = ConnectionInfo[ConnInfoInd['line_id']]

		# add only a single line for a series of connections on the same line
		if not line == LineList[-1]:
			LineList.append(line_id)
	return LineSet

def GetNumberOfLineChanges(PathInfo, ConnectionInfo=None):
	"""
	Get total number of line changes within the path.
	On-foot connections are not counted as line changes. 
	"""
	# shortcut
	if len(PathInfo) <= 2: return 0 

	# on-foot gattung
	OnFootGattungList = TrWay.values()

	# count line changes 
	LineChanges = 0 
	for i in range(2, len(PathInfo)):
		if PathInfo[i][ConnInfoInd['line_category']] not in OnFootGattungList and PathInfo[i-1][ConnInfoInd['travel_id']] != PathInfo[i][ConnInfoInd['travel_id']]:
			LineChanges += 1
	if ConnectionInfo and ConnectionInfo[ConnInfoInd['line_category']] not in OnFootGattungList and ConnectionInfo[ConnInfoInd['travel_id']] != PathInfo[-1][ConnInfoInd['travel_id']]:
		LineChanges += 1
	return LineChanges

def GetNumberOfOnFootConnections(PathInfo):
	"""
	Get the total number of on-foot (zu Fuss) connections in the path.
	OnFoot connection means, gattung in TrWay.values()
	"""
	NumberOfOnFootConnections = 0 
	OnFootGattungList = TrWay.values()

	for ConnInfo in PathInfo:
		gattung = ConnInfo[ConnInfoInd['line_category']]
		if gattung in OnFootGattungList:
			NumberOfOnFootConnections += 1 
	return NumberOfOnFootConnections

def GetMeasurableLinesPerTimeWindow(PathInfo, TimeWindows, RequiredMeasureTime):
	"""
	Get possible line measurements per each time window (Zeitfenster).
	Normally, TimeWindows = ZF, like {1: (0, 5*60+59), 2: (6*60, 10*60+59), ...}
	Output example:
	MeasurementPerTimeWindow = {
			(11,'S','12'): [1,2],			# line can be measured in time windows 1 and 2
			(11,'S','15'): [3],
		}
	NOTE:
	How many minutes can a measurement continue into following time window?
	RequiredMeasureTime -1? Is it decisive at which line the measurement has started?
	"""
	MeasurementPerTimeWindow = {}
	if PathInfo == None: return None
	if len(PathInfo) < 2: return None

	AllLinesOfPath = GetAllLinesOfPath(PathInfo)
	TimeIntervalsForEachLine = GetTimeIntervalsForEachLine(PathInfo)

	for Line in AllLinesOfPath:
		for key in TimeWindows:
			TimeWindow = TimeWindows[key]
			MeasureIntervalBegin = TimeWindow[0]
			MeasureIntervalEnd = TimeWindow[1] + (RequiredMeasureTime - 1) 	# check assumption !!!

			TimeIntv = GetTimeIntervalForMeasurementOnLine(Line, TimeIntervalsForEachLine, \
				MeasureIntervalBegin, MeasureIntervalEnd, RequiredMeasureTime)
			if TimeIntv:
				if MeasurementPerTimeWindow.has_key(Line):
					MeasurementPerTimeWindow[Line].append(key)
				else:
					MeasurementPerTimeWindow[Line] = [key]
	return MeasurementPerTimeWindow

def GetMeasurableStationsPerTimeWindow(PathInfo, TimeWindows, RequiredMeasureTime):
	"""
	Get possible station measurements per each time window (Zeitfenster).
	Normally, TimeWindows = ZF, like {1: (0, 5*60+59), 2: (6*60, 10*60+59), ...}
	Output example:
	MeasurementPerTimeWindow = {
			s1: [1,2],			# station can be measured in time windows 1 and 2
			s2: [3],
		}
	NOTE:
	How many minutes can a measurement continue into following time window?
	RequiredMeasureTime -1? Is it decisive at which line the measurement has started?
	"""
	MeasurementPerTimeWindow = {}
	if PathInfo == None: return None
	if len(PathInfo) < 2: return None

	AllStationsOfPath = GetAllStationsOfRoute(PathInfo)
	TimeIntervalsForEachStation = GetTimeIntervalsForEachStation(PathInfo)

	for Station in AllStationsOfPath:
		for key in TimeWindows:
			TimeWindow = TimeWindows[key]
			MeasureIntervalBegin = TimeWindow[0]
			MeasureIntervalEnd = TimeWindow[1] + (RequiredMeasureTime - 1) 	# check assumption !!!

			TimeIntv = GetTimeIntervalForMeasurementAtStation(Station, TimeIntervalsForEachStation, \
				MeasureIntervalBegin, MeasureIntervalEnd, RequiredMeasureTime)
			
			if TimeIntv:
				if MeasurementPerTimeWindow.has_key(Station):
					MeasurementPerTimeWindow[Station].append(key)
				else:
					MeasurementPerTimeWindow[Station] = [key]
	return MeasurementPerTimeWindow

# START measurement/interval evaluation

def GetTimeIntervalsForEachLine(PathInfo):
	"""
	Return journey time spent for each line.
	Note: Ignores first connection (ConnectionInfo) in PathInfo
	Returns TimeIntervalOnLine is a dictionary like:
	TimeIntervalOnLine[LineID] = 
		[(FirstDepartureTimeMin1,LastArrivalTimeMin1),(FirstDepartureTimeMin2,LastArrivalTimeMin2)]
	"""
	TimeIntervalOnLine = {}

	if PathInfo == None: return TimeIntervalOnLine
	if len(PathInfo) < 2: return TimeIntervalOnLine

	LastLine = None
	CurrDeparture = None

	for i in range(1, len(PathInfo)):
		ConnectionInfo = PathInfo[i]
		CurLine = ConnectionInfo[ConnInfoInd['line_id']]
		CurTripID = ConnectionInfo[ConnInfoInd['travel_id']]

		if i == 1:
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID

		# if CurLine != LastLine:
		if CurTripID != LastTripID:
			LastArrival = PathInfo[i-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[i-1][ConnInfoInd['arrival_min']]
			if TimeIntervalOnLine.has_key(LastLine):
				TimeIntervalOnLine[LastLine].append((CurrDeparture, LastArrival))
			else:
				TimeIntervalOnLine[LastLine] = [(CurrDeparture, LastArrival)]
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID
		
		if i == (len(PathInfo) - 1):
			LastArrival = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 + ConnectionInfo[ConnInfoInd['arrival_min']]
			if TimeIntervalOnLine.has_key(CurLine):
				TimeIntervalOnLine[CurLine].append((CurrDeparture, LastArrival))
			else:
				TimeIntervalOnLine[CurLine] = [(CurrDeparture, LastArrival)]

	return TimeIntervalOnLine

def GetTimeIntervalsForEachStation(PathInfo):
	"""
	Return time spent at each station in minutes. 
	Returned TimeIntervalAtStation is a dictionary like:
	TimeIntervalAtStation[station] = [(ArrivalTimeMin1,DepartureTimeMin1),(ArrivalTimeMin2,DepartureTimeMin2)]
	NOTE: A station can be visited multiple times 
	unless VisitAStationOnlyOnce condition is given.
	"""
	TimeIntervalAtStation = {}

	if not PathInfo: return TimeIntervalAtStation
	if len(PathInfo) < 2: return TimeIntervalAtStation

	for i in range(1, len(PathInfo)):
		Connection1 = PathInfo[i-1]
		Connection2 = PathInfo[i]

		TripID1 = Connection1[ConnInfoInd['travel_id']]
		TripID2 = Connection2[ConnInfoInd['travel_id']]

		# check if customer makes a change at station
		# if TripID1 == TripID2: continue

		station = Connection2[ConnInfoInd['station_from']]

		ArrivalMin = Connection1[ConnInfoInd['arrival_hour']]*60 + Connection1[ConnInfoInd['arrival_min']]
		DepartureMin = Connection2[ConnInfoInd['departure_hour']]*60 + Connection2[ConnInfoInd['departure_min']]

		if TimeIntervalAtStation.has_key(station):
			TimeIntervalAtStation[station].append((ArrivalMin, DepartureMin))
		else:
			TimeIntervalAtStation[station] = [(ArrivalMin, DepartureMin)]
	return TimeIntervalAtStation

def GetTimeIntervalsOfMeasurableStations(PathInfo, RequiredMeasureTime, ExcludedStationList, RequiredLinesForMeasurement):
	"""
	Get time intervals (IntervalInfo) for each measurable station, together with line-information:
	To which line is the station-measurement assigned.

	A station is measurable, if:
	1) There is enough time at the station for a measurement (time at station >= RequiredStationMeasureTime)
	2) station is not in ExcludedStationList
	3) directly following line (departure line) is in RequiredLinesForMeasurement; that is, a LinienID in project.

	A station-measurement is normally assigned to directly following departure line (and line --> line bundle).
	Exception: The last station of a tour is assigned to its arrival line.

	TimeIntervalsOfMeasurableStations[station] = [(Line1, IntvStartMin1, IntvEndMin1), (Line2, IntvStartMin2, IntvEndMin2), ...]
	"""
	TimeIntervalsOfMeasurableStations = {}

	if not PathInfo: return TimeIntervalsOfMeasurableStations
	if len(PathInfo) < 2: return TimeIntervalsOfMeasurableStations

	for i in range(1, len(PathInfo)):
		Connection1 = PathInfo[i-1]
		Connection2 = PathInfo[i]
		station = Connection2[ConnInfoInd['station_from']]

		TripID1 = Connection1[ConnInfoInd['travel_id']]
		TripID2 = Connection2[ConnInfoInd['travel_id']]

		# check if customer makes a change at station
		# if TripID1 == TripID2: continue

		# check if station should be measured
		if station in ExcludedStationList: continue

		# check if there is enough time for a measurement
		ArrivalMin = Connection1[ConnInfoInd['arrival_hour']]*60 + Connection1[ConnInfoInd['arrival_min']]
		DepartureMin = Connection2[ConnInfoInd['departure_hour']]*60 + Connection2[ConnInfoInd['departure_min']]
		if (DepartureMin-ArrivalMin) < RequiredMeasureTime: continue

		# check if assigned line should be measured
		line = Connection2[ConnInfoInd['line_id']]
		if line not in RequiredLinesForMeasurement: continue

		# add IntervalInfo to dictionary 
		IntervalInfo = (line, ArrivalMin, DepartureMin)
		if not TimeIntervalsOfMeasurableStations.has_key(station): TimeIntervalsOfMeasurableStations[station] = []
		TimeIntervalsOfMeasurableStations[station].append(IntervalInfo)

	# check last station of tour
	FirstStation = PathInfo[0][ConnInfoInd['station_to']]
	LastStation = PathInfo[-1][ConnInfoInd['station_to']]

	# if round tour --> first station (= last) must normally be measured at the beginning
	if not TimeIntervalsOfMeasurableStations.has_key(LastStation):
		# check arrival line
		Connection2 = PathInfo[-1]

		ArrivalMin = Connection2[ConnInfoInd['arrival_hour']]*60 + Connection1[ConnInfoInd['arrival_min']]
		
		line = Connection2[ConnInfoInd['line_id']]

		if line in RequiredLinesForMeasurement and LastStation not in ExcludedStationList:
			TimeIntervalsOfMeasurableStations[LastStation] = [(line, ArrivalMin, ArrivalMin+RequiredMeasureTime)]

	return TimeIntervalsOfMeasurableStations

def GetTimeIntervalsOfMeasurableStations2(PathInfo, RequiredMeasureTime):
	"""
	Get time intervals (IntervalInfo) for each measurable station, together with line-information:
	To which line is the station-measurement assigned.

	A station is measurable, if:
	1) There is enough time at the station for a measurement (time at station >= RequiredStationMeasureTime)

	A station-measurement is normally assigned to directly following departure line (and line --> line bundle).
	Exception: The last station of a tour is assigned to its arrival line.

	TimeIntervalsOfMeasurableStations[station] = [(Line1, IntvStartMin1, IntvEndMin1), (Line2, IntvStartMin2, IntvEndMin2), ...]
	"""
	TimeIntervalsOfMeasurableStations = {}

	if not PathInfo: return TimeIntervalsOfMeasurableStations
	if len(PathInfo) < 2: return TimeIntervalsOfMeasurableStations

	for i in range(1, len(PathInfo)):
		Connection1 = PathInfo[i-1]
		Connection2 = PathInfo[i]
		station = Connection2[ConnInfoInd['station_from']]

		TripID1 = Connection1[ConnInfoInd['travel_id']]
		TripID2 = Connection2[ConnInfoInd['travel_id']]

		# check if customer makes a change at station
		# if TripID1 == TripID2: continue

		# check if there is enough time for a measurement
		ArrivalMin = Connection1[ConnInfoInd['arrival_hour']]*60 + Connection1[ConnInfoInd['arrival_min']]
		DepartureMin = Connection2[ConnInfoInd['departure_hour']]*60 + Connection2[ConnInfoInd['departure_min']]
		if (DepartureMin-ArrivalMin) < RequiredMeasureTime: continue

		# check if assigned line should be measured
		line = Connection2[ConnInfoInd['line_id']]

		# add IntervalInfo to dictionary 
		IntervalInfo = (line, ArrivalMin, DepartureMin)
		if not TimeIntervalsOfMeasurableStations.has_key(station): TimeIntervalsOfMeasurableStations[station] = []
		TimeIntervalsOfMeasurableStations[station].append(IntervalInfo)

	# check last station of tour
	FirstStation = PathInfo[0][ConnInfoInd['station_to']]
	LastStation = PathInfo[-1][ConnInfoInd['station_to']]

	# if round tour --> first station (= last) must normally be measured at the beginning
	if not TimeIntervalsOfMeasurableStations.has_key(LastStation):
		# check arrival line
		Connection2 = PathInfo[-1]

		ArrivalMin = Connection2[ConnInfoInd['arrival_hour']]*60 + Connection1[ConnInfoInd['arrival_min']]

		line = Connection2[ConnInfoInd['line_id']]

		TimeIntervalsOfMeasurableStations[LastStation] = [(line, ArrivalMin, ArrivalMin+RequiredMeasureTime)]

	return TimeIntervalsOfMeasurableStations

def GetMeasurableLines(PathInfo, TimeWindows, WeekDayGroup, RequiredMeasureTime, StartDate, EndDate):
	"""
	Get all possible line measurements per time window (Zeitfenster)
	and weekday group (Wochentagtyp) within the given time interval.
	Returns MeasurableLines = {
		((11, 'S', '7'), 1, 11): "",
		((11, 'S', '9'), 2, 12): "",
	}
	Note: LineMeasurementRequirements = {
		((11, 'S', '7'), 1, 11): 24,
		((11, 'S', '7'), 1, 12): 26,
		((11, 'S', '9'), 2, 12): 35,
	}
	"""
	# get measurable lines per time window
	MeasurementPerTimeWindow = GetMeasurableLinesPerTimeWindow(PathInfo, TimeWindows, RequiredMeasureTime)

	# get available weekdays
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(PathInfo, StartDate, EndDate)
	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysRoute)

	TimeWindows = ZF 
	WeekdayGroups = WD 

	# get included weekday groups in AvailableWeekDays
	IncludedWeekdayGroups = []
	AvlWeekDaysSet = set(AvailableWeekDays)
	
	for key in WeekdayGroups:
		if set(WeekdayGroups[key]).intersection(AvlWeekDaysSet):
			IncludedWeekdayGroups.append(key)

	MeasurableLines = {}

	for line in MeasurementPerTimeWindow:
		for TimeWindowKey in MeasurementPerTimeWindow[line]:
			for WeekdayGroupKey in IncludedWeekdayGroups:
				MeasurableLines[line, TimeWindowKey, WeekdayGroupKey] = ""
	return MeasurableLines

def GetMeasurableStations(PathInfo, TimeWindows, WeekDayGroup, RequiredMeasureTime, StartDate, EndDate):
	"""
	Get all possible station measurements per time window (Zeitfenster)
	and weekday group (Wochentagtyp) within the given time interval.
	Returns MeasurableStations = {
		(8503340, 1, 11): "",
		(8503309, 2, 12): "",
	}
	Note: StationMeasurementRequirements = {
		(8503340, 1, 11): 24,
		(8503309, 1, 12): 26,
		(8503308, 2, 12): 35,
	}
	"""
	# get measurable stations per time window
	MeasurementPerTimeWindow = GetMeasurableStationsPerTimeWindow(PathInfo, TimeWindows, RequiredMeasureTime)

	# get available weekdays
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(PathInfo, StartDate, EndDate)
	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysRoute)

	TimeWindows = ZF 
	WeekdayGroups = WD 

	# get included weekday groups in AvailableWeekDays
	IncludedWeekdayGroups = []
	AvlWeekDaysSet = set(AvailableWeekDays)
	
	for key in WeekdayGroups:
		if set(WeekdayGroups[key]).intersection(AvlWeekDaysSet):
			IncludedWeekdayGroups.append(key)

	MeasurableStations = {}

	for station in MeasurementPerTimeWindow:
		for TimeWindowKey in MeasurementPerTimeWindow[station]:
			for WeekdayGroupKey in IncludedWeekdayGroups:
				MeasurableStations[station, TimeWindowKey, WeekdayGroupKey] = ""
	return MeasurableStations

def GetMeasurableTimeIntervalsPerLineKey(PathInfo, MinMeasureTime, ZF, WD, StartDate, EndDate):
	"""
	Get measurable time intervals per line, together with FahrtID information.

	Return for each line (lineID) time interval spent (riding time)
	and FahrtID, provided that SpentTime >= MinMeasureTime (in minutes),
	within the given time interval.
	Note: Ignores first connection (ConnectionInfo) in PathInfo
	
	Returns TimeIntervalPerLine is a dictionary like:
	TimeIntervalPerLine[(line, Zeitfenster, Wochentagtyp)] = [
		(FahrtID1, DepartureMin1,LastArrivalMin1), 
		(FahrtID2, DepartureMin2,LastArrivalMin2),
	]
	"""
	TimeIntervalPerLine = {}

	if PathInfo == None: return TimeIntervalPerLine
	if len(PathInfo) < 2: return TimeIntervalPerLine

	# get available weekdays
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(PathInfo, StartDate, EndDate)
	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysRoute)

	TimeWindows = ZF 
	WeekdayGroups = WD 

	# get included weekday groups in AvailableWeekDays
	IncludedWeekdayGroups = []
	AvlWeekDaysSet = set(AvailableWeekDays)
	
	for key in WeekdayGroups:
		if set(WeekdayGroups[key]).intersection(AvlWeekDaysSet):
			IncludedWeekdayGroups.append(key)

	LastLine = None
	CurrDeparture = None

	for i in range(1, len(PathInfo)):
		ConnectionInfo = PathInfo[i]
		CurLine = ConnectionInfo[ConnInfoInd['line_id']]
		CurTripID = ConnectionInfo[ConnInfoInd['travel_id']]

		if i == 1:
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID

		# if CurLine != LastLine:
		if CurTripID != LastTripID:
			LastArrival = PathInfo[i-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[i-1][ConnInfoInd['arrival_min']]
			MatchingTW = GetMatchingTimeWindows(ZF, CurrDeparture, LastArrival, MinMeasureTime)

			if MatchingTW:
				for tw in MatchingTW:
					for wdg in IncludedWeekdayGroups:
						linekey = (LastLine,tw,wdg)
						if TimeIntervalPerLine.has_key(linekey):
							TimeIntervalPerLine[linekey].append((LastTripID, CurrDeparture, LastArrival))
						else:
							TimeIntervalPerLine[linekey] = [(LastTripID, CurrDeparture, LastArrival)]
			
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID
		
		if i == (len(PathInfo) - 1):
			LastArrival = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 + ConnectionInfo[ConnInfoInd['arrival_min']]
			MatchingTW = GetMatchingTimeWindows(ZF, CurrDeparture, LastArrival, MinMeasureTime)

			if MatchingTW:
				for tw in MatchingTW:
					for wdg in IncludedWeekdayGroups:
						linekey = (LastLine,tw,wdg)
						if TimeIntervalPerLine.has_key(linekey):
							TimeIntervalPerLine[linekey].append((CurTripID, CurrDeparture, LastArrival))
						else:
							TimeIntervalPerLine[linekey] = [(CurTripID, CurrDeparture, LastArrival)]

	return TimeIntervalPerLine

def GetMeasurableTimeIntervalsPerStationKey(PathInfo, MinMeasureTime, ZF, WD, StartDate, EndDate):
	"""
	Get measurable time intervals per station, together with FahrtID information.

	Return for each station time interval (waiting time) spent 
	provided that SpentTime >= MinMeasureTime (in minutes),
	within the given time interval.
	Note: Ignores first connection (ConnectionInfo) in PathInfo
	
	Returns TimeIntervalPerStation is a dictionary like:
	TimeIntervalPerStation[(station, Zeitfenster, Wochentagtyp)] = [
		(DepartureMin1,LastArrivalMin1), 
		(DepartureMin2,LastArrivalMin2),
	]
	"""
	TimeIntervalPerStation = {}
	if PathInfo == None: return TimeIntervalPerStation
	if len(PathInfo) < 2: return TimeIntervalPerStation

	# get available weekdays
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(PathInfo, StartDate, EndDate)
	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysRoute)

	TimeWindows = ZF 
	WeekdayGroups = WD 

	# get included weekday groups in AvailableWeekDays
	IncludedWeekdayGroups = []
	AvlWeekDaysSet = set(AvailableWeekDays)
	
	for key in WeekdayGroups:
		if set(WeekdayGroups[key]).intersection(AvlWeekDaysSet):
			IncludedWeekdayGroups.append(key)

	# TimeIntervalAtStation[station] = [(ArrivalTimeMin1,DepartureTimeMin1),(ArrivalTimeMin2,DepartureTimeMin2)]
	TimeIntervalAtStation = GetTimeIntervalsForEachStation(PathInfo)

	for station in TimeIntervalAtStation:
		for interval in TimeIntervalAtStation[station]:
			ArrivalTimeMin = interval[0]
			DepartureTimeMin = interval[1]
			if (DepartureTimeMin - ArrivalTimeMin) >= MinMeasureTime:
				MatchedTimeWindows = GetMatchingTimeWindows(ZF, ArrivalTimeMin, DepartureTimeMin, MinMeasureTime)
				for wdgroup in IncludedWeekdayGroups:
					for twindow in MatchedTimeWindows:
						statkey = (station, twindow, wdgroup)
						if TimeIntervalPerStation.has_key(statkey):
							TimeIntervalPerStation[statkey].append((ArrivalTimeMin, DepartureTimeMin))
						else:
							TimeIntervalPerStation[statkey] = [(ArrivalTimeMin, DepartureTimeMin)]
	return TimeIntervalPerStatio

def GetMeasurableTimeIntervalsPerLineAndTW(RouteInfo, MinMeasureTime, ZF):
	"""
	Get measurable time intervals per line and TimeWindow (ZF), together with FahrtID information.

	Return for each Line (lineID) and TimeWindow, time interval spent (riding time)
	and FahrtID, provided that SpentTime >= MinMeasureTime (in minutes),
	within the given time interval.
	Note: Ignores first connection (ConnectionInfo) in PathInfo
	
	Returns TimeIntervalPerLine is a dictionary like:
	TimeIntervalPerLineAndZF[(Line, TimeWindow)] = [
		(FahrtID1, DepartureMin1,LastArrivalMin1), 
		(FahrtID2, DepartureMin2,LastArrivalMin2),
	]
	"""
	TimeIntervalPerLineAndTW = {}
	LastLine = None
	CurrDeparture = None

	for i in range(1, len(RouteInfo)):
		ConnectionInfo = RouteInfo[i]
		CurLine = ConnectionInfo[ConnInfoInd['line_id']]
		CurTripID = ConnectionInfo[ConnInfoInd['travel_id']]

		if i == 1:
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID

		# if CurLine != LastLine:
		if CurTripID != LastTripID:
			LastArrival = RouteInfo[i-1][ConnInfoInd['arrival_hour']]*60 + RouteInfo[i-1][ConnInfoInd['arrival_min']]
			MatchingTW = GetMatchingTimeWindows(ZF, CurrDeparture, LastArrival, MinMeasureTime)

			if MatchingTW:
				for tw in MatchingTW:
					linekey = (LastLine,tw)
					if TimeIntervalPerLineAndTW.has_key(linekey):
						TimeIntervalPerLineAndTW[linekey].append((LastTripID, CurrDeparture, LastArrival))
					else:
						TimeIntervalPerLineAndTW[linekey] = [(LastTripID, CurrDeparture, LastArrival)]
			
			CurrDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
			LastLine = CurLine
			LastTripID = CurTripID
		
		if i == (len(RouteInfo) - 1):
			LastArrival = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 + ConnectionInfo[ConnInfoInd['arrival_min']]
			MatchingTW = GetMatchingTimeWindows(ZF, CurrDeparture, LastArrival, MinMeasureTime)

			if MatchingTW:
				for tw in MatchingTW:
						linekey = (LastLine,tw)
						if TimeIntervalPerLineAndTW.has_key(linekey):
							TimeIntervalPerLineAndTW[linekey].append((CurTripID, CurrDeparture, LastArrival))
						else:
							TimeIntervalPerLineAndTW[linekey] = [(CurTripID, CurrDeparture, LastArrival)]
	return TimeIntervalPerLineAndTW

def GetRequiredMeasurableTimeIntervalsPerLineAndTW(RouteInfo, MinMeasureTime, ZF, WD, StartDate, EndDate, LineMeasurementRequirements):
	"""
	Get measurable time intervals per Line and TimeWindow (ZF) 2-tuples
	that are required, together with FahrtID information.

	Return for each Line (lineID) and TimeWindow, time interval spent (riding time)
	and FahrtID, provided that SpentTime >= MinMeasureTime (in minutes),
	within the given time interval.
	Note: Ignores first connection (ConnectionInfo) in RouteInfo

	LineMeasurementRequirements = {
		((11, 'S', '7'), 2, 12): 2,			# ZF[2]: (6:00-10:59), WD[11]: (1,2,3,4,5) d.h. Arbeitstage
		((11, 'S', '7'), 3, 11): 15,
		...
	}
	TimeIntervalPerLineAndZF[(Line, TimeWindow)] = [
		(FahrtID1, DepartureMin1,LastArrivalMin1), 
		(FahrtID2, DepartureMin2,LastArrivalMin2),
	]
	Returns ReqTimeIntervalPerLineAndZF = filtered TimeIntervalPerLineAndZF
	"""
	# shortcut 
	if LineMeasurementRequirements == None: return {}
	if len(LineMeasurementRequirements) == 0: return {}
	TimeIntervalPerLineAndTW = GetMeasurableTimeIntervalsPerLineAndTW(RouteInfo, MinMeasureTime, ZF)

	# get availability of route
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(RouteInfo, StartDate, EndDate)
	WeekdayGroupsOfRoute = GetAvailableWeekDayGroups(WD, AvailableDaysRoute)

	# compare measurable lines with requirements --> filter
	RequiredTimeIntvPerLineAndZF = {}
	for LineTWkey in TimeIntervalPerLineAndTW:
		Line = LineTWkey[0]
		TW = LineTWkey[1]
		for wdg in WeekdayGroupsOfRoute:
			key = (Line, TW, wdg)
			if key in LineMeasurementRequirements:
				RequiredTimeIntvPerLineAndZF[LineTWkey] = TimeIntervalPerLineAndTW[LineTWkey]
				break
	return RequiredTimeIntvPerLineAndZF

def CheckIfEquivalentLMProfiles(LMCoveragePerLineKey1, LMCoveragePerLineKey2, LineMeasurementReq=None):
	"""
	Return True, if Line Measurement (LM) Coverage of two routes/variants are identical.

	LMCoveragePerLineKey[(Line, TW, WG)] = n
	"""
	ConsiderLMRequirements = False 
	if LineMeasurementReq:
		ConsiderLMRequirements = True
	
	# get relevant line keys
	RelevantLineKeys = []
	if ConsiderLMRequirements:
		for LineKey in LineMeasurementReq:
			if LineMeasurementReq[LineKey] > 0:
				RelevantLineKeys.append(LineKey)
	
	if ConsiderLMRequirements:
		LMCoveragePerLineKey1 = SelectKeysFromDic(LMCoveragePerLineKey1, RelevantLineKeys)
		LMCoveragePerLineKey2 = SelectKeysFromDic(LMCoveragePerLineKey2, RelevantLineKeys)
	
	return CheckIfEqualDic(LMCoveragePerLineKey1, LMCoveragePerLineKey2, DefaultEmptyVal = 0)

def CheckIfLMProfileRoute1ContainsLMProfileRoute2(LMCoveragePerLineKey1, LMCoveragePerLineKey2, LineMeasurementReq=None):
	"""
	Return True, if Line Measurement (LM) Coverage of the first route contains the LM Coverage of the second route.

	Contains means:
	- LMCoveragePerLineKey1 has all the LineKeys of LMCoveragePerLineKey2, and
	- LMCoveragePerLineKey1[LineKey] >= LMCoveragePerLineKey2[LineKey] for all LineKeys.

	LMCoveragePerLineKey[(Line, TW, WG)] = n
	"""
	ConsiderLMRequirements = False 
	if LineMeasurementReq:
		ConsiderLMRequirements = True
	
	# get relevant line keys
	RelevantLineKeys = []
	if ConsiderLMRequirements:
		for LineKey in LineMeasurementReq:
			if LineMeasurementReq[LineKey] > 0:
				RelevantLineKeys.append(LineKey)
	
	if ConsiderLMRequirements:
		LMCoveragePerLineKey1 = SelectKeysFromDic(LMCoveragePerLineKey1, RelevantLineKeys)
		LMCoveragePerLineKey2 = SelectKeysFromDic(LMCoveragePerLineKey2, RelevantLineKeys)

	# check containment
	for LineKey in LMCoveragePerLineKey2:
		if not LineKey in LMCoveragePerLineKey1:
			return False 
		if LMCoveragePerLineKey1[LineKey] < LMCoveragePerLineKey2[LineKey]:
			return False 
	
	return True

def SelectBestRoutesForLineMeasurement(SortedRouteInfoList, MultiplicityLimit, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, 
	LineMeasurementReq=None):
	"""
	Select best routes, that add to Line Measurement Coverage, up to MultiplicityLimit for a LineKey (line, TW, WG).
	A route with lower value than an already selected equivalent route is ignored; i.e. is not selected.
	
	SortedRouteInfoList: After value sorted routes, in descending order.
	MultiplicityLimit: Upper limit (like 3) for added values per LineKey. 
		No new tour that adds to a LineKey is selected, after MultiplicityLimit for this LineKey is achieved.

	Returns: (SelectedRoutes, TotalLMCoverage)
			TotalLMCoverage: Total LM coverage of selected routes

	Required globals:
	- TimeWindows
	- WeekDayGroups
	"""
	# init
	SelectedRoutes = []
	TotalLMCoverage = {}
	LMCoverageOfSelectedRoutes = []

	for RouteInfo in SortedRouteInfoList:
		(LMCoverageOfRoutePerSegment, LMCoverageOfRoutePerLineKey) = \
			GetLMCoverageOfRoute(RouteInfo, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LineMeasurementReq)

		# ignore LM equivalent or contained routes
		# obsolete: does nothing, waste of time
		"""
		for LMCov in LMCoverageOfSelectedRoutes:
			# if CheckIfEquivalentLMProfiles(LMCov, LMCoverageOfRoutePerLineKey):
			if CheckIfLMProfileRoute1ContainsLMProfileRoute2(LMCov, LMCoverageOfRoutePerLineKey):
				# don't select route, continue with the next route
				continue
		"""

		# check MultiplicityLimit and added value
		# added value: Measurement of a LineKey for which TotalLMCoverage[LineKey] < MultiplicityLimit
		for LineKey in LMCoverageOfRoutePerLineKey:
			if LMCoverageOfRoutePerLineKey[LineKey] == 0:
				continue
			if LineKey not in TotalLMCoverage.keys():
				TotalLMCoverage = AddDicValues(TotalLMCoverage, LMCoverageOfRoutePerLineKey, DefaultEmptyVal = 0)
				SelectedRoutes.append(RouteInfo)
				LMCoverageOfSelectedRoutes.append(LMCoverageOfRoutePerLineKey)
				break
			elif TotalLMCoverage[LineKey] < MultiplicityLimit:
				TotalLMCoverage = AddDicValues(TotalLMCoverage, LMCoverageOfRoutePerLineKey, DefaultEmptyVal = 0)
				SelectedRoutes.append(RouteInfo)
				LMCoverageOfSelectedRoutes.append(LMCoverageOfRoutePerLineKey)
				break
			else: 
				pass
	return (SelectedRoutes, TotalLMCoverage)

def AddBestValueRoutesToSelectedRoutes(SelectedRoutes, NewRoutes, MultiplicityLimit, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, 
	RouteValueFunc, SortParameters, LineMeasurementReq=None):
	"""
	Add new routes to a list of selected routes considering route value and LM profile of route.

	Replace a route in SelectedRoutes with a new route, if the LM profile of the new route contains the other's, 
	but its value is higher.

	MultiplicityLimit: Upper limit (like 3) for added values per LineKey. 
		No new tour that adds to a LineKey is selected, after MultiplicityLimit for this LineKey is achieved.

	Returns: (SelectedRoutes, TotalLMCoverage)
			TotalLMCoverage: Total LM coverage of selected routes

	Required globals:
	- TimeWindows
	- WeekDayGroups
	"""
	# combine routes
	AllRoutes = SelectedRoutes + NewRoutes

	# sort all routes
	SortedRoutes = SortRoutesAfterValueInDescOrder(AllRoutes, RouteValueFunc, SortParameters)

	# select best routes
	return SelectBestRoutesForLineMeasurement(SortedRoutes, MultiplicityLimit, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LineMeasurementReq)

def GetRouteSegments(RouteInfo, TimeWindows):
	"""
	Get extended route segment information with corresponding time windows.

	A line trip (Reise/Travel with a unique trip id) is splitted into multiple segments 
	if it spans multiple time windows or LineIDs. 

	The departure time from a station determines the time window of the segment.
	Each route segment corresponds to a unique TimeWindow index (1-5).

	Virtual line-change points: Whenever TimeWindow or LineID change

	A route segment is measurable, if the time interval between  
		a) first departure (i.e. departure from the first station of segment), and
		b) final arrival time of the line 
	is larger than ReqLineMeasureTime.

	Returns RouteSegments (dictionary), where RouteSegments[SegmentNr] = SegmentInfo,
	where SegmentNr starts from 1, not 0.
	
	SegmentInfo: A list whose content is defined with SegmentInfoInd.
	"""
	RouteSegments = {}

	if not RouteInfo:
		return RouteSegments 
	if len(RouteInfo) < 2:
		return RouteSegments

	SegmentNr = 0
	LastTripID = None
	LastTimeWindow = None
	LastLineID = None
	SegmentsOfCurrentLine = set()

	for i in range(1, len(RouteInfo)):
		Connection = RouteInfo[i]

		TripID = Connection[ConnInfoInd['travel_id']]
		LineID = Connection[ConnInfoInd['line_id']]
		CurrDepartureTime = 60* Connection[ConnInfoInd['departure_hour']] + Connection[ConnInfoInd['departure_min']]
		CurrTimeWindowOfSegment = FindTimeWindowOfTimePoint(ZF, CurrDepartureTime)

		# line change
		if i == 1 or (TripID != LastTripID or LineID != LastLineID) or TripID == None:	
			SegmentNr += 1 
			LastTripID = TripID
			LastLineID = LineID 
			
			RouteSegments[SegmentNr] = [None] * len(SegmentInfoInd)
			RouteSegments[SegmentNr][SegmentInfoInd['FirstConnectionInd']] = i 
			RouteSegments[SegmentNr][SegmentInfoInd['TimeWindow']] = CurrTimeWindowOfSegment
			LastTimeWindow = CurrTimeWindowOfSegment

			RouteSegments[SegmentNr][SegmentInfoInd['trip_id']] = TripID
			RouteSegments[SegmentNr][SegmentInfoInd['first_station']] = Connection[ConnInfoInd['station_from']]
			RouteSegments[SegmentNr][SegmentInfoInd['line_id']] = Connection[ConnInfoInd['line_id']]
			RouteSegments[SegmentNr][SegmentInfoInd['linie']] = Connection[ConnInfoInd['line']]
			RouteSegments[SegmentNr][SegmentInfoInd['gattung']] = Connection[ConnInfoInd['line_category']]
			RouteSegments[SegmentNr][SegmentInfoInd['verwaltung']] = Connection[ConnInfoInd['management']]
			RouteSegments[SegmentNr][SegmentInfoInd['fahrtnum']] = Connection[ConnInfoInd['travel_no']]
			RouteSegments[SegmentNr][SegmentInfoInd['line_IntvStart']] = CurrDepartureTime
			
			if i == 1:
				PrevConnection = RouteInfo[i-1]
				ArrivalTimeOfPreviousLine = 60 *PrevConnection[ConnInfoInd['arrival_hour']] + PrevConnection[ConnInfoInd['arrival_min']]
				
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvStart']] = ArrivalTimeOfPreviousLine
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvEnd']] = CurrDepartureTime
				
				SegmentsOfCurrentLine.add(SegmentNr)

			# TripID changes --> new segment
			if i > 1:
				PrevConnection = RouteInfo[i-1]
				ArrivalTimeOfPreviousLine = 60 *PrevConnection[ConnInfoInd['arrival_hour']] + PrevConnection[ConnInfoInd['arrival_min']]
				RouteSegments[SegmentNr-1][SegmentInfoInd['line_IntvEnd']] = ArrivalTimeOfPreviousLine
				RouteSegments[SegmentNr-1][SegmentInfoInd['last_station']] = PrevConnection[ConnInfoInd['station_to']]
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvStart']] = RouteSegments[SegmentNr-1][SegmentInfoInd['line_IntvEnd']]
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvEnd']] = CurrDepartureTime
				
				for seg in SegmentsOfCurrentLine:
					RouteSegments[seg][SegmentInfoInd['FinalArrivalTimeOfLine']] = ArrivalTimeOfPreviousLine
				
				SegmentsOfCurrentLine = set()
				SegmentsOfCurrentLine.add(SegmentNr)

		# line does not change
		else:
			# same line, new segment due to new TimeWindow
			if LastTimeWindow != CurrTimeWindowOfSegment:
				SegmentNr += 1 

				RouteSegments[SegmentNr] = [None] * len(SegmentInfoInd)
				RouteSegments[SegmentNr][SegmentInfoInd['FirstConnectionInd']] = i 
				RouteSegments[SegmentNr][SegmentInfoInd['TimeWindow']] = CurrTimeWindowOfSegment
				LastTimeWindow = CurrTimeWindowOfSegment
				SegmentsOfCurrentLine.add(SegmentNr)

				RouteSegments[SegmentNr][SegmentInfoInd['trip_id']] = TripID
				RouteSegments[SegmentNr][SegmentInfoInd['first_station']] = Connection[ConnInfoInd['station_from']]
				RouteSegments[SegmentNr][SegmentInfoInd['line_id']] = Connection[ConnInfoInd['line_id']]
				RouteSegments[SegmentNr][SegmentInfoInd['linie']] = Connection[ConnInfoInd['line']]
				RouteSegments[SegmentNr][SegmentInfoInd['gattung']] = Connection[ConnInfoInd['line_category']]
				RouteSegments[SegmentNr][SegmentInfoInd['verwaltung']] = Connection[ConnInfoInd['management']]
				RouteSegments[SegmentNr][SegmentInfoInd['fahrtnum']] = Connection[ConnInfoInd['travel_no']]
				RouteSegments[SegmentNr][SegmentInfoInd['line_IntvStart']] = 60 *Connection[ConnInfoInd['departure_hour']] + Connection[ConnInfoInd['departure_min']]
				
				PrevConnection = RouteInfo[i-1]
				RouteSegments[SegmentNr-1][SegmentInfoInd['line_IntvEnd']] = 60 *PrevConnection[ConnInfoInd['arrival_hour']] + PrevConnection[ConnInfoInd['arrival_min']]
				RouteSegments[SegmentNr-1][SegmentInfoInd['last_station']] = PrevConnection[ConnInfoInd['station_to']]
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvStart']] = RouteSegments[SegmentNr-1][SegmentInfoInd['line_IntvEnd']]
				RouteSegments[SegmentNr][SegmentInfoInd['stat_IntvEnd']] = RouteSegments[SegmentNr][SegmentInfoInd['line_IntvStart']]

			# same line, same TimeWindow, segment does not change
			else:
				pass

		# last connection of route
		if i == (len(RouteInfo) -1):	
			FinalArrivalTimeOfLine = 60 *Connection[ConnInfoInd['arrival_hour']] + Connection[ConnInfoInd['arrival_min']]
			RouteSegments[SegmentNr][SegmentInfoInd['line_IntvEnd']] = FinalArrivalTimeOfLine
			RouteSegments[SegmentNr][SegmentInfoInd['last_station']] = Connection[ConnInfoInd['station_to']]
			
			for seg in SegmentsOfCurrentLine:
				RouteSegments[seg][SegmentInfoInd['FinalArrivalTimeOfLine']] = FinalArrivalTimeOfLine
				
			# check if there is sufficient time to measure the last station
			# if RouteSegments[SegmentNr][SegmentInfoInd['line_IntvEnd']] - RouteSegments[1][SegmentInfoInd['stat_IntvStart']] + TimeMarginAfterRouteEnds <= MaxTourDuration:

	return RouteSegments

def GetTravelSegments(RouteInfo, TimeWindows):
	"""
	Get travel segments (Reisen) together with the TimeWindows and LineIDs they span.

	The departure time from a station determines the time window of a possible 
	line measurement.

	This function calls: GetRouteSegments()

	Returns TravelSegments (dictionary), where TravelSegments[SegmentNr] = TravelInfo,
	where SegmentNr starts from 1, not 0.
	
	TravelInfo: A dictionary whose content is defined with SegmentInfoInd.
	"""
	TravelSegments = {}

	if not RouteInfo:
		return TravelSegments 
	if len(RouteInfo) < 2:
		return TravelSegments

	# get route segments with virtual change points
	RouteSegments = GetRouteSegments(RouteInfo, TimeWindows)

	PreviousTripID = None
	TravelNr = 0

	for SegmentNr in RouteSegments:
		SegmentInfo = RouteSegments[SegmentNr]
		TripID = SegmentInfo[SegmentInfoInd['trip_id']]
	
		if TripID == None or TripID != PreviousTripID:
			# line change: new line# line change: new line	

			TravelNr += 1 
			TravelInfo = [None] * len(SegmentInfoInd)
			PreviousTripID = TripID

			# Same TravelInfo over changing route segments (virtual change points)
			TravelInfo[SegmentInfoInd['FirstConnectionInd']] = SegmentInfo[SegmentInfoInd['FirstConnectionInd']]
			TravelInfo[SegmentInfoInd['trip_id']] = TripID
			TravelInfo[SegmentInfoInd['first_station']] = SegmentInfo[SegmentInfoInd['first_station']]
			TravelInfo[SegmentInfoInd['verwaltung']] = SegmentInfo[SegmentInfoInd['verwaltung']]
			TravelInfo[SegmentInfoInd['gattung']] = SegmentInfo[SegmentInfoInd['gattung']]
			TravelInfo[SegmentInfoInd['linie']] = SegmentInfo[SegmentInfoInd['linie']]
			TravelInfo[SegmentInfoInd['fahrtnum']] = SegmentInfo[SegmentInfoInd['fahrtnum']]
			TravelInfo[SegmentInfoInd['line_IntvStart']] = SegmentInfo[SegmentInfoInd['line_IntvStart']]
			TravelInfo[SegmentInfoInd['stat_IntvStart']] = SegmentInfo[SegmentInfoInd['stat_IntvStart']]
			TravelInfo[SegmentInfoInd['stat_IntvEnd']] = SegmentInfo[SegmentInfoInd['stat_IntvEnd']]
			TravelInfo[SegmentInfoInd['FinalArrivalTimeOfLine']] = SegmentInfo[SegmentInfoInd['FinalArrivalTimeOfLine']]
			
			# NOT same TravelInfo over changing route segments (virtual change points)
			TravelInfo[SegmentInfoInd['last_station']] = SegmentInfo[SegmentInfoInd['last_station']]
			TravelInfo[SegmentInfoInd['line_IntvEnd']] = SegmentInfo[SegmentInfoInd['line_IntvEnd']]

			# ADD to list
			TravelInfo[SegmentInfoInd['TimeWindow']] = [ SegmentInfo[SegmentInfoInd['TimeWindow']] ]
			TravelInfo[SegmentInfoInd['line_id']] = [ SegmentInfo[SegmentInfoInd['line_id']] ]

			TravelSegments[TravelNr] = TravelInfo

		else:	
			# no line change
			TravelInfo = TravelSegments[TravelNr]

			# NOT same TravelInfo over changing route segments (virtual change points)
			TravelInfo[SegmentInfoInd['last_station']] = SegmentInfo[SegmentInfoInd['last_station']]
			TravelInfo[SegmentInfoInd['line_IntvEnd']] = SegmentInfo[SegmentInfoInd['line_IntvEnd']]

			# ADD to list
			TimeWindow = SegmentInfo[SegmentInfoInd['TimeWindow']] 
			if TimeWindow not in TravelInfo[SegmentInfoInd['TimeWindow']]:
				TravelInfo[SegmentInfoInd['TimeWindow']].append(TimeWindow)

			LineID = SegmentInfo[SegmentInfoInd['line_id']]
			if LineID not in TravelInfo[SegmentInfoInd['line_id']]:
				TravelInfo[SegmentInfoInd['line_id']].append(LineID)

	return TravelSegments

def GetLineMeasurementCoverageOfRoute(RouteInfo, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LMRequirements=None, ExcludeTW0=True, ExcludeWG10=True):
	"""
	Which lines a route can measure: Line Measurement (LM) Coverage (Umfang) per Route Segment.
	Evaluate LM coverage based on (a) segments and (b) available weekdays of route.

	Note: Only lines with LineID values can be measured (LineID not in [-1, None])

	ExcludeTW0: If True, ignore (don't include) TimeWindow 0 in outputs.
	ExcludeWG10: If True, ignore (don't include) WeekdayGroup 10 in outputs.
	LMRequirements: If None (or empty), ignore requirements to calculate coverage. If not None,
		calculate coverage w.r.t. LMRequirements.
	LMRequirements[(LineID, TW, WG)] = x

	Required global variables:
	ZF: Zeitfenster 
	WD: Weekday Groups

	Functions to generate input parameters:
	- (AvailableDaysOfRoute, UnavailableDaysOfRoute) = GetAvailabilityOfRoute(RouteInfo, FirstDayOfPeriod, LastDayOfPeriod)
	- RouteSegments = GetRouteSegments(RouteInfo)

	Note: The line of a segment is measurable, if the time interval between 
	the first departure time (of segment) and the final arrival time of line 
	>= ReqLineMeasureTime
	
	Returns:
	1) LMCoverageOfRoutePerSegment[SegmentNr] = (LineID, 2, FahrtID, LineIntvStart, LineIntvEnd)
		.. where 2 is the measurable time windows of the segment.
		Only measurable segments are included in the dictionary.

	2) LMCoverageOfRoutePerLineKey[(Line, TW, WG)] = n
		.. where n = 1, 2, 3 ... multiplicity of a LineKey measurement.
		A LineID can be measurable multiple times within a route.
		LineKey: (LineID, TimeWindow, WeekdayGroup) = (Line, TW, WG)

	LM rule: A FahrtID can be measured only once within a route.
	"""
	if RouteInfo == None: return None 

	# get route segments
	(AvailableDaysOfRoute, UnavailableDaysOfRoute) = GetAvailabilityOfRoute(RouteInfo, FirstDayOfPeriod, LastDayOfPeriod)
	RouteSegments = GetRouteSegments(RouteInfo, ZF)

	ConsiderLMRequirements = False
	if LMRequirements: 
		ConsiderLMRequirements = True

	# LMRequirements: TW/WG sets 
	TimeWindowsPerLine = {}
	WeekdayGroupsPerLineAndTW = {}
	WGroups = set()
	if ConsiderLMRequirements:
		for LineKey in LMRequirements:
			(line, tw, wg) = LineKey
			if not TimeWindowsPerLine.has_key(line): TimeWindowsPerLine[line] = set()
			TimeWindowsPerLine[line].add(tw)
			if line not in WeekdayGroupsPerLineAndTW: WeekdayGroupsPerLineAndTW[(line,tw)] = set()
			WeekdayGroupsPerLineAndTW[(line,tw)].add(wg)
			WGroups.add(wg)

	# LineKey: (LineID, TimeWindow, WeekdayGroup) = (Line, TW, WG)
	LMCoverageOfRoutePerLineKey = {}
	LMCoverageOfRoutePerSegment = {}

	if not AvailableDaysOfRoute: 
		return (LMCoverageOfRoutePerSegment, LMCoverageOfRoutePerLineKey)

	# get available weekdays
	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysOfRoute)

	TimeWindows = ZF 
	WG = WD 

	# get included weekday groups in AvailableWeekDays
	IncludedWeekdayGroups = []
	AvlWeekDaysSet = set(AvailableWeekDays)
	
	for key in WG:
		if set(WG[key]).intersection(AvlWeekDaysSet):
			IncludedWeekdayGroups.append(key)

	if ExcludeWG10:
		IncludedWeekdayGroups.remove(10)

	if ConsiderLMRequirements:
		IncludedWeekdayGroups = WGroups.intersection(set(IncludedWeekdayGroups))

	# evaluate each segment of route
	for SegmentNr in RouteSegments.keys():
		SegmentInfo = RouteSegments[SegmentNr]
		LineID = SegmentInfo[SegmentInfoInd['line_id']]
		TimeWindow = SegmentInfo[SegmentInfoInd['TimeWindow']]
		
		if not LineID or LineID == "-1": 
			continue 
		
		# if ConsiderLMRequirements and not TimeWindowsPerLine.has_key(LineID):
		if ConsiderLMRequirements and (not CheckIfLineIDAndTimeWindowInKey(LMRequirements, LineID, TimeWindow) \
			or not WeekdayGroupsPerLineAndTW[(LineID,TimeWindow)].intersection(IncludedWeekdayGroups)):
			continue

		FahrtID = SegmentInfo[SegmentInfoInd['trip_id']]
		LineIntvStart = SegmentInfo[SegmentInfoInd['line_IntvStart']]
		# LineIntvEnd = SegmentInfo[SegmentInfoInd['line_IntvEnd']]
		LineFinalEnd = SegmentInfo[SegmentInfoInd['FinalArrivalTimeOfLine']]
		

		# check if line of segment is measurable
		if LineFinalEnd - LineIntvStart >= ReqLineMeasureTime:
			LMCoverageOfRoutePerSegment[SegmentNr] = (LineID, TimeWindow, FahrtID, LineIntvStart, LineFinalEnd)
		
			# add to LMCoverageOfRoutePerLineKey
			for wdg in IncludedWeekdayGroups:
				LineKey = (LineID,TimeWindow,wdg)
				
				# 7.3.2017: Tunc
				if ConsiderLMRequirements and not LineKey in LMRequirements:
					continue
				
				if LMCoverageOfRoutePerLineKey.has_key(LineKey):
					LMCoverageOfRoutePerLineKey[LineKey] += 1
				else:
					LMCoverageOfRoutePerLineKey[LineKey] = 1 

	return (LMCoverageOfRoutePerSegment, LMCoverageOfRoutePerLineKey)

GetLMCoverageOfRoute = GetLineMeasurementCoverageOfRoute 	# alias name

# new function: 8. March 2020 by Tunc
def GetLMCoverageOfRouteForGivenDay(RouteInfo, DayOrd, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LMRequirements=None, ExcludeTW0=True, ExcludeWG10=True):
	"""
	Get concrete Line Measurement (LM) Coverage of a route for a given DayOrd like date(2018, 4, 10).toordinal()
	"""
	(LMCoverageOfRoutePerSegment, PotentialLMCoverage) = GetLMCoverageOfRoute(RouteInfo, ReqLineMeasureTime, FirstDayOfPeriod, 
		LastDayOfPeriod, LMRequirements, ExcludeTW0, ExcludeWG10)
	ConcreteLMCoverage = {}

	if not PotentialLMCoverage:
		return ConcreteLMCoverage

	# get WeekdayGroup of day
	WeekdayGroups = GetWeekdayGroupsOfDate(WD, DayOrd)
	wg = WeekdayGroups[0]

	for (LineID, TWindow, WGroup) in PotentialLMCoverage:
		if WGroup == wg:
			ConcreteLMCoverage[(LineID, TWindow, WGroup)] = 1

	return ConcreteLMCoverage

# updated: 11.05.2017 (error corrected)
def GetLMCoverageOfMultipleRoutes(RouteInfoList, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LMRequirements=None):
	"""
	Get Line Measurement (LM) Coverage of all routes in the list (RouteList).

	If LMRequirements is not None, only the required (relevant) line keys (Line, TW, WG)
	of LMRequirements are considered for the coverage.

	Returns:
	LMCoverageOfRoutesPerLineKey[(Line, TW, WG)] = n
		where n is the multiplicity of LM coverage.
	"""
	LMCoverageOfRoutesPerLineKey = {}

	for RouteInfo in RouteInfoList:
		(LMCoverageOfRoutePerSegment, LMCoverageOfRoutePerLineKey) = \
			GetLMCoverageOfRoute(RouteInfo, ReqLineMeasureTime, FirstDayOfPeriod, LastDayOfPeriod, LMRequirements)

		# add dictionaries
		LMCoverageOfRoutesPerLineKey = AddDicValues(LMCoverageOfRoutesPerLineKey, LMCoverageOfRoutePerLineKey, DefaultEmptyVal = 0)

	return LMCoverageOfRoutesPerLineKey

def GetStationMeasurementCoverageOfRoute(RouteInfo, AQMeasureTime, KIMeasureTime, AQMeasureTimePerStation, 
	FirstDayOfPeriod, LastDayOfPeriod, ReturnType=1, LineToReqBundle=None):
	"""
	Which stations a route can measure: Station Measurement (SM) Coverage (Umfang) per Route Segment,
	for both AQ and KI type station measurements.
	Evaluate SM coverage based on extended route segments.

	Note: Only stations with corresponding valid LineID values can be measured (LineID not in [-1, None])

	Return seperate measurability results (dictionaries) for:
	- AQ only 
	- KI only 
	- Both AQ and KI

	ReturnType: 
		1: Return per LineID results 
		2: Return per Station results 
		3: Return per LineBundle results (mapping by LineToReqBundle)

	AQMeasureTimePerStation[station] = StationSpecificMeasureTimeAQ

	LineToReqBundle: If None (or empty), ignore SM requirements to calculate coverage. 
		If not None, calculate coverage w.r.t. LineToReqBundle.
		Note: Only required (for measurements) line bundles are included in the 
		dictionary LineToReqBundle as values.
	
	Functions to generate input parameters:
	- (AvailableDaysOfRoute, UnavailableDaysOfRoute) = GetAvailabilityOfRoute(RouteInfo, FirstDayOfPeriod, LastDayOfPeriod)
	- RouteSegments = GetRouteSegments(RouteInfo)
	
	Returns:
	1) SMCoverageOfRoutePerSegment[SegmentNr] = (Station, LineID, MCode, IntvStart, IntvEnd)
		MCode: Measurability code 
		1: Can measure AQ only
		2: Can measure KI only 
		3: Can measure either AQ or KI, but not both 
		4: Can measure both AQ and KI
	2) SMCoverageAQonly[key] = n
		.. where n = 1, 2, 3 ... multiplicity of a station measurement.
		A station can be measurable multiple times within a route.
		key can be LineID, Station or LineBundle depending on ReturnType.
	3) SMCoverageKIonly[key] = n
	4) SMCoverageBothAQandKI[key] = n
		
	Notes: 
	* Both AQ and KI type station measurements must be mapped to a LineBundle over LineID,
		if LineToBundle is given. 
	* LineID assigment for a segment must be handled by func GetRouteSegments

	Returns:
	(SMCoverageOfRoutePerSegment, SMCoverageAQonly, SMCoverageKIonly, SMCoverageBoth)
	"""
	if RouteInfo == None: return None 

	# check availability of route	
	(AvailableDaysOfRoute, UnavailableDaysOfRoute) = GetAvailabilityOfRoute(RouteInfo, FirstDayOfPeriod, LastDayOfPeriod)
	if not AvailableDaysOfRoute: 
		return ({}, {}, {}, {})

	# get route segments
	RouteSegments = GetRouteSegments(RouteInfo, ZF)
	SMCoverageOfRoutePerSegment = {}

	# SM coverage per LineID, Station or LineBundle (depending on ReturnType)
	SMCoverageAQonly = {}
	SMCoverageKIonly = {}
	SMCoverageBoth = {}			# both AQ and KI

	ConsiderSMRequirements = False
	if LineToReqBundle:
		ConsiderSMRequirements = True

	# return type
	ReturnPerLineID = False 
	ReturnPerStation = False 
	ReturnPerLineBundle = False 
	if ReturnType == 1:
		ReturnPerLineID = True 
	elif ReturnType == 2:
		ReturnPerStation = True 
	elif ReturnType == 3:
		ReturnPerLineBundle = True
	else:
		raise Exception("Undefined ReturnType %s!" % ReturnType)

	# evaluate each segment of route
	for SegmentNr in RouteSegments.keys():
		SegmentInfo = RouteSegments[SegmentNr]
		LineID = SegmentInfo[SegmentInfoInd['line_id']]
		Gattung = SegmentInfo[SegmentInfoInd['gattung']]
		
		# is LineID messrelevant?
		IfRelevantLine = LineID not in [None, '-1'] and (not ConsiderSMRequirements or LineID in LineToReqBundle)

		# is W-type last segment for AQ measurement only?
		IfLastWSegmentForAQ = SegmentNr == len(RouteSegments) and Gattung == 'W'

		station = SegmentInfo[SegmentInfoInd['first_station']]
		StatIntvStart = SegmentInfo[SegmentInfoInd['stat_IntvStart']]
		StatIntvEnd = SegmentInfo[SegmentInfoInd['stat_IntvEnd']]

		MCode = 0

		# get station-specific measure time
		AQMeasureTimeX = AQMeasureTime
		if AQMeasureTimePerStation and station in AQMeasureTimePerStation:
			AQMeasureTimeX = AQMeasureTimePerStation[station]
		
		# AQ only
		if IfRelevantLine and StatIntvEnd - StatIntvStart >= AQMeasureTimeX:
			MCode += 1
			
			if ReturnPerLineID:
				IncrementDicValue(SMCoverageAQonly, LineID)
			elif ReturnPerStation:
				IncrementDicValue(SMCoverageAQonly, station)
			else: 
				if LineToReqBundle and LineID in LineToReqBundle:
					LineBundle = LineToReqBundle[LineID]
					IncrementDicValue(SMCoverageAQonly, LineBundle)
		
		# KI only
		if IfRelevantLine and StatIntvEnd - StatIntvStart >= KIMeasureTime and not IfLastWSegmentForAQ:
			MCode += 2
			
			if ReturnPerLineID:
				IncrementDicValue(SMCoverageKIonly, LineID)
			elif ReturnPerStation:
				IncrementDicValue(SMCoverageKIonly, station)
			else: 
				if LineToReqBundle and LineID in LineToReqBundle:
					LineBundle = LineToReqBundle[LineID]
					IncrementDicValue(SMCoverageKIonly, LineBundle)
		
		# Both AQ and KI
		if IfRelevantLine and StatIntvEnd - StatIntvStart >= AQMeasureTimeX + KIMeasureTime and not IfLastWSegmentForAQ:
			MCode += 1
			
			if ReturnPerLineID:
				IncrementDicValue(SMCoverageBoth, LineID)
			elif ReturnPerStation:
				IncrementDicValue(SMCoverageBoth, station)
			else: 
				if LineToReqBundle and LineID in LineToReqBundle:
					LineBundle = LineToReqBundle[LineID]
					IncrementDicValue(SMCoverageBoth, LineBundle)

		if MCode > 0:
			SMCoverageOfRoutePerSegment[SegmentNr] = (station, LineID, MCode, StatIntvStart, StatIntvEnd)

	return (SMCoverageOfRoutePerSegment, SMCoverageAQonly, SMCoverageKIonly, SMCoverageBoth)

GetSMCoverageOfRoute = GetStationMeasurementCoverageOfRoute 

def GetExactTrailOfRoute(RouteInfo):
	"""
	Get exact trail information of route to return following lists:

	StationList: List of stations to be visited 
	IntervalBeginTimes: List of interval begin times in total minutes, sorted in ascending order 
	IntervalEndTimes: 	List of interval end times in total minutes, sorted in ascending order 

	Note: Time intervals are given with respect to arrival times (ankunft).
	See related TourCondition FollowTrail_with_ExclusiveTimeIntervals
	"""
	ArtificialStation = 8500000 
	StationList = []
	IntervalBeginTimes = []
	IntervalEndTimes = []

	for ConnInfo in RouteInfo:
		departure_station = ConnInfo[ConnInfoInd['station_from']]
		if departure_station == ArtificialStation:
			continue 
		arrival_station = ConnInfo[ConnInfoInd['station_to']]
		arrival_time = ConnInfo[ConnInfoInd['arrival_hour']]*60 + ConnInfo[ConnInfoInd['arrival_min']]

		StationList.append(arrival_station)
		IntervalBeginTimes.append(arrival_time)
		IntervalEndTimes.append(arrival_time)

	return (StationList, IntervalBeginTimes, IntervalEndTimes)

# Route Value (see also GetRouteValue in AssignmentFunctions)

def GetSimpleValueOfRoute(PathInfo, MeasurementTime=10, MeasurementValue=100, HourlyCost=40):
	"""
	Calculate monetary value of a route (assuming all measurements are required) considering:
	MeasurementTime: Minimum time in minutes required for a measurement on a line
	MeasurementValue: Monatary value of a single measurement
	HourlyCost: Hourly cost of a trip
	Note: A line can be measured at most once in a journey.
	"""
	MeasuredLine = []
	if PathInfo == None: return 0
	if len(PathInfo) < 2: return 0

	TimeIntervals = GetTimeIntervalsForEachLine(PathInfo)
	MeasurementCount = 0

	# find all possible measurements
	for Line in TimeIntervals.keys():
		if not Line in MeasuredLine:
			intervals = TimeIntervals[Line]
			for interval in intervals:
				DepartureTime = interval[0]
				ArrivalTime = interval[1] 
				if  (ArrivalTime - DepartureTime) >= MeasurementTime:
					MeasurementCount +=1
					MeasuredLine.append(Line)
					break 
	# total trip time in minutes
	TripStartTime = PathInfo[1][ConnInfoInd['abfahrt_std']]*60 + PathInfo[1][ConnInfoInd['abfahrt_min']]
	TripEndTime = PathInfo[-1][ConnInfoInd['ankunft_std']]*60 + PathInfo[-1][ConnInfoInd['ankunft_min']]
	TotalTripDuration = TripEndTime - TripStartTime

	# trip value
	TripValue = MeasurementValue * MeasurementCount - (TotalTripDuration / 60.0 * HourlyCost)
	return TripValue

def GetSimpleValueOfRoute(PathInfo, TimeWindows, WeekDayGroup, ReqLineMeasureTime, ReqStationMeasureTime, \
	StartDate, EndDate, HourlyTripCost, IncomeLineMeasure, IncomeStationMeasure, \
	LineMeasurementReq, LineToBundle):
	"""
	Evaluate the value of route with respect to its measurable lines and stations within the 
	given time range.

	Assumption to simplify: A line or station can be measured only once within a route.

	TimeIntervalsOfMeasurableStations[station] = [(Line1, IntvStartMin1, IntvEndMin1), (Line2, IntvStartMin2, IntvEndMin2), ...]
	"""
	MeasurableLines = GetMeasurableLines(PathInfo, TimeWindows, WeekDayGroup, ReqLineMeasureTime, StartDate, EndDate)
	# MeasurableStations = GetMeasurableStations(PathInfo, TimeWindows, WeekDayGroup, ReqStationMeasureTime, StartDate, EndDate)
	TimeIntervalsOfMeasurableStations = GetTimeIntervalsOfMeasurableStations2(PathInfo, ReqStationMeasureTime)

	# get duration of route
	TotalDuration = GetTotalDurationOfPath(PathInfo)

	# how many lines can be measured
	LineRegister = []	# to ensure that a line is measured only once
	LineMeasureCount = 0

	for key in LineMeasurementReq:
		line = key[0]
		if MeasurableLines.has_key(key) and line not in LineRegister:
			LineMeasureCount += 1
			LineRegister.append(line)

	# how many stations can be measured
	StationRegister = []	# to ensure that a station is measured only once
	StationMeasureCount = 0

	for station in TimeIntervalsOfMeasurableStations:
		if station in StationRegister:
			continue
		intervals = TimeIntervalsOfMeasurableStations[station]
		for interval in intervals:
			(line,smin,emin) = interval 
			if LineToBundle.has_key(line):
				StationMeasureCount += 1 
				StationRegister.append(station)
				break

	RouteValue = LineMeasureCount*IncomeLineMeasure + StationMeasureCount*IncomeStationMeasure \
		- (TotalDuration / 60.0) * HourlyTripCost
	return RouteValue

def SortRoutesAfterValueInDescOrder(RouteInfoList, RouteValueFunc, Parameters):
	"""
	Sort routes after their values, in descending order.
	Return a sorted route list.

	RouteValueFunc: Name of function, that evaluates the value of a route.
	Parameters: n-Tuple with parameters for RouteValueFunc.
		(param1, param2, param3, ...)
	"""
	RouteInd = range(0, len(RouteInfoList))
	RouteValues = []

	for i in RouteInd:
		RouteInfo = RouteInfoList[i]
		RouteValue = RouteValueFunc(RouteInfo, *Parameters)
		RouteValues.append(RouteValue)

	SortedInd = SortIndex(RouteValues, RouteInd)
	SortedInd.reverse()	
	
	SortedRouteInfoList = []
	for i in SortedInd:
		SortedRouteInfoList.append(RouteInfoList[i])
	return SortedRouteInfoList


# **************************************************************************************
# Availability and Weekday functions (Verkehrstage und Werktage)
# **************************************************************************************

def GetAvailabilityBetweenDates(StartDate, EndDate, VerkehrstageHex):
	"""
	Return available and unavailable dates as ordinals (date.toordinal).
	StartDate and EndDate: dates like date(2008,12,5)
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	# Get Last Day of the Month in Python
	# http://stackoverflow.com/questions/42950/get-last-day-of-the-month-in-python

	# reference dates as ordinals
	FPLANStartDayOrd = FPLAN_BeginDate.toordinal()
	FPLANEndDayOrd = FPLAN_EndDate.toordinal()

	# dates as ordinal numbers
	StartDateOrd = StartDate.toordinal()
	EndDateOrd = EndDate.toordinal()

	# check given dates
	if StartDateOrd > FPLANEndDayOrd:
		raise Exception("StartDate cannot be larger (more recent) than EndDate!")
	if StartDateOrd < FPLANStartDayOrd or StartDateOrd > FPLANEndDayOrd:
		raise Exception("StartDate must lie between FPLAN start and end dates!")
	if EndDateOrd < FPLANStartDayOrd or EndDateOrd > FPLANEndDayOrd:
		raise Exception("EndDate must lie between FPLAN start and end dates!")

	AvailableDaysOrd = []
	UnavailableDaysOrd = []

	# check if VerkehrstageHex is null or empty
	if VerkehrstageHex == '' or VerkehrstageHex == None:
		for DateOrd in range(StartDateOrd, EndDateOrd+1):
			ref = DateOrd - FPLANStartDayOrd
			AvailableDaysOrd.append(DateOrd)
		return (AvailableDaysOrd, UnavailableDaysOrd)

	# convert hexa to bin
	VerkehrstageBin = hextobin(VerkehrstageHex)

	for DateOrd in range(StartDateOrd, EndDateOrd+1):
		ref = DateOrd - FPLANStartDayOrd
		if VerkehrstageBin[2 + ref] == '1':
			AvailableDaysOrd.append(DateOrd)
		else:
			UnavailableDaysOrd.append(DateOrd)
	return (AvailableDaysOrd, UnavailableDaysOrd)

def ConvertAvailableDayListToHexCode(DayOrdList, StartDate=FPLAN_BeginDate, EndDate=FPLAN_EndDate):
	"""
	Convert a list of available days (DayOrdList) to hexcode (VerkehrstageHex)
	"""
	# reference dates as ordinals
	StartDateOrd = StartDate.toordinal()
	EndDayOrd = EndDate.toordinal()

	BinStr = '11'

	for DayOrd in range(StartDateOrd, EndDayOrd):
		BinD = '0'
		if DayOrd in DayOrdList:
			BinD = '1'
		BinStr += BinD 

	BinStr += '11'
	print BinStr
	return bintohex(BinStr)

def GetAvailabilityForMonth(year, month, VerkehrstageHex):
	"""
	Return available and unavailable dates as ordinals (date.toordinal).
	Inputs: year like 2016, month like 12
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	# Get Last Day of the Month in Python
	# http://stackoverflow.com/questions/42950/get-last-day-of-the-month-in-python
	"""
	monthrange(year, month):
    Returns weekday of first day of the month and number of days in month
    """
	DayRange = calendar.monthrange(year,month)
	MonthEndDay = DayRange[1]

	StartDate = date(year,month,1)
	EndDate = date(year,month,MonthEndDay)		
	return GetAvailabilityBetweenDates(StartDate, EndDate, VerkehrstageHex)

def GetAvailabilityOfConnection(ConnectionInfo, StartDate, EndDate):
	"""
	Return available and unavailable dates for connection as ordinals (date.toordinal)
	between StartDate and EndDate: dates like date(2008,12,5).
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	VerkehrstageHex = ConnectionInfo[ConnInfoInd['trafficdays_hexcode']]
	return GetAvailabilityBetweenDates(StartDate, EndDate, VerkehrstageHex)

def GetAvailabilityOfConnectionForMonth(ConnectionInfo, year, month):
	"""
	Return available and unavailable dates as ordinals (date.toordinal) for connection.
	Inputs: year like 2016, month like 12
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	VerkehrstageHex = ConnectionInfo[ConnInfoInd['trafficdays_hexcode']]
	return GetAvailabilityForMonth(year, month, VerkehrstageHex)

def GetAvailabilityOfRoute(RouteInfo, StartDate, EndDate):
	"""
	Return available and unavailable dates for route as ordinals (date.toordinal)
	between StartDate and EndDate: dates like date(2008,12,5).
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	AvailableDaysRoute = set()
	UnavailableDaysRoute = set()

	ctr = 0
	for ConnInfo in RouteInfo:
		ctr += 1
		VerkehrstageHex = ConnInfo[ConnInfoInd['trafficdays_hexcode']]
		(AvailableDays, UnavailableDays) = GetAvailabilityBetweenDates(StartDate, EndDate, VerkehrstageHex)
		
		if ctr == 1:
			AvailableDaysRoute = set(AvailableDays)
		else:
			AvailableDaysRoute = AvailableDaysRoute.intersection(set(AvailableDays))

		UnavailableDaysRoute = UnavailableDaysRoute.union(set(UnavailableDays))

	AvailableDaysRoute = list(AvailableDaysRoute)
	UnavailableDaysRoute = list(UnavailableDaysRoute)
	
	AvailableDaysRoute.sort()
	UnavailableDaysRoute.sort()
	return (AvailableDaysRoute, UnavailableDaysRoute)

def GetAvailabilityOfRouteForMonth(RouteInfo, year, month):
	"""
	Return available and unavailable dates as ordinals (date.toordinal) for route.
	Inputs: year like 2016, month like 12
	Returns date lists as ordinals: (AvailableDaysOrd, UnavailableDaysOrd)
	"""
	DayRange = calendar.monthrange(year,month)
	MonthEndDay = DayRange[1]

	StartDate = date(year,month,1)
	EndDate = date(year,month,MonthEndDay)	
	return GetAvailabilityOfRoute(RouteInfo, StartDate, EndDate)

def GetAvailableWeekDays(AvailableDaysOrd):
	"""
	Obtain available weekdays from the list of available days, including ordinal dates.
	Return an availability list like [1,2,3,4,5] 
	where 1 is for Monday, 7 is for Sunday.
	"""
	if AvailableDaysOrd == None: return []

	AvailableWeekDays = set()

	for DayOrd in AvailableDaysOrd:
		WeekDay = date.fromordinal(DayOrd).isoweekday()
		AvailableWeekDays.add(WeekDay)
		if len(AvailableWeekDays) == 7: 
			awd = list(AvailableWeekDays)
			awd.sort()
			return awd
	awd = list(AvailableWeekDays)
	awd.sort()
	return awd

def GetAvailableWeekDayGroups(WD, AvailableDaysOrd):
	"""
	Obtain available WeekDayGroups (as defined in WD) from the list of available days
	including ordinal dates.
	Return an WD availability list like [10,11] 
	"""
	if AvailableDaysOrd == None: return []

	AvailableWeekDays = GetAvailableWeekDays(AvailableDaysOrd)
	AvailableWeekdayGroups = set()

	for WeekDay in AvailableWeekDays:
		for WDkey in WD:
			if WeekDay in WD[WDkey]:
				AvailableWeekdayGroups.add(WDkey)
		if len(AvailableWeekdayGroups) == len(WD): 
			wdg = list(AvailableWeekdayGroups)
			wdg.sort()
			return wdg
	wdg = list(AvailableWeekdayGroups)
	wdg.sort()
	return wdg

# **************************************************************************************
# Condition functions that return condition string for reading db table
# **************************************************************************************

def CondStrRouteStartAndDuration(StartHour, StartMin, LatestArrivalIn):
	"""
	Return SQL condition string for route start and max duration in minutes.

	LatestArrivalIn: Time interval from RouteStart to latest arrival.
	"""
	# MarginMin = 30
	# MarginMin = MaxWaitingTime
	StartMin = StartHour*60 + StartMin
	
	# SQLcond = " abfahrtm >= %s AND abfahrtm <= %s " % \
	#	(StartMin-MarginMin, StartMin+LatestArrivalIn+MarginMin)
	
	SQLcond = " departure_totalmin >= %s AND departure_totalmin <= %s AND arrival_totalmin <= %s " % \
		(StartMin, StartMin+LatestArrivalIn, StartMin+LatestArrivalIn)

	return SQLcond

def CondStrSelectedWeekDays(WeekDayList):
	"""
	Return SQL condition string for selected week days.
	example: WeekDayList = [6,7] --> find all connections that go on both 6. and 7. days of the week
	1 means Monday, 7 is Sunday
	"""
	if not WeekDayList:
		return ""

	# see: 9.7. Pattern Matching, similar to
	# http://www.postgresql.org/docs/8.3/static/functions-matching.html
	# select '1|2|3' similar to '%(3|4|5)%'

	WeekDayList.sort()			# sort list
	wds = set(WeekDayList)		# eliminate recurring numbers

	# SQLcond = "wochentage SIMILAR TO " + \
	#	"'%(" + "|".join([str(e) for e in wds]) + ")%'"

	SQLcond = "wochentage LIKE " + "'%" + "%".join([str(e) for e in wds]) + "%'"

	return SQLcond

def CondStrSelectedGattung(GattungList):
	"""
	Return SQL condition string for Gattung which must be included in GattungList.
	example: GattungList = ['S', 'RE', 'BUS']
	"""
	# SQLcond = "gattung SIMILAR TO " + "'(" + "|".join(GattungList) + ")'"

	condstr = []
	for gattung in GattungList:
		pat = "'" + gattung + "'"
		condstr.append(pat)
	SQLcond = "line_category IN " + "(" + ",".join(condstr) + ")"
	return SQLcond

def CondStrSelectedStations(StationList):
	"""
	Return SQL condition string for selected stations (Include-Only stations).
	example: StationList = [8503000, 8503006, 8503129, 8503306, 8503147]
	"""
	vals = "(" + ",".join([str(e) for e in StationList]) + ")"
	SQLcond = "station_from in " + vals + " AND " + "station_to in " + vals
	return SQLcond

def CondStrSelectedLines(LineList):
	"""
	Return SQL condition string for selected lines (Include-Only lines).
	line --> linie_id like '110|S|9'
	example: LineList = ['110|S|9', '110|S|11', '110|S|12', ...]
	"""
	condstr = []
	for line in LineList:
		pat = "'" + line + "'"
		condstr.append(pat)
	SQLcond = "line_id IN " + "(" + ",".join(condstr) + ")"
	return SQLcond

def CondStrSelectedTripDays_hex(DayOrdList):
	"""
	Return SQL condition string for selected trip days.
	example: DayList = ['20-12-2015', '15-2-2016']
	Selected connections must be available on ALL listed days.

	works with hex code (verkehrstage_hexcode)
	"""
	# 9.4. String Functions and Operators
	# http://www.postgresql.org/docs/9.4/static/functions-string.html
	# http://stackoverflow.com/questions/17208945/whats-the-easiest-way-to-represent-a-bytea-as-a-single-integer-in-postgresql

	# start day reference
	StartDayRef = FPLAN_BeginDate.toordinal()
	EndDayRef = FPLAN_EndDate.toordinal()

	# convert date string to day count w.r.t. reference day
	DayRefList = []
	for day in DayOrdList:
		ref = day - StartDayRef 

		if ref < 0 or ref > (EndDayRef - StartDayRef):
			raise Exception("All dates in DayList must be between FPLAN START and END days!")
		DayRefList.append(ref)

	# eliminate recurring dates
	DayRefSet = set(DayRefList)

	condstr = []
	for ref in DayRefSet:
		refi = ref + 3
		# select ('x' || verkehrstage_hexcode)::bit(240) 
		cstr = "substring((('x' || trafficdays_hexcode)::bit(" + str(refi) + "))::text from " + str(refi) + " for 1) = '1'"
		condstr.append(cstr)
	SQLcond = " (trafficdays_hexcode is null OR \n(" + "\n AND ".join(condstr) + "))\n "

	return SQLcond

def CondStrSelectedTripDays_bitfeld(DayList):
	"""
	Return SQL condition string for selected trip days.
	example: DayList = ['20-12-2015', '15-2-2016']
	Selected connections must be available on ALL listed days.

	works with hex code (verkehrstage_bitfeld)
	"""
	# 9.4. String Functions and Operators
	# http://www.postgresql.org/docs/9.4/static/functions-string.html
	# http://stackoverflow.com/questions/17208945/whats-the-easiest-way-to-represent-a-bytea-as-a-single-integer-in-postgresql

	# start day reference
	StartDayRef = FPLAN_BeginDate.toordinal()
	EndDayRef = FPLAN_EndDate.toordinal()

	# convert date string to day count w.r.t. reference day
	DayRefList = []
	for dstr in DayList:
		day = datetime.strptime(dstr, "%d-%m-%Y").date()
		ref = day.toordinal() - StartDayRef 

		if ref < 0 or ref > (EndDayRef - StartDayRef):
			raise Exception("All dates in DayList must be between FPLAN START and END days!")
		DayRefList.append(ref)

	# eliminate recurring dates
	DayRefSet = set(DayRefList)

	condstr = []
	for ref in DayRefSet:
		refi = ref + 1 
		cstr = "substring(trafficdays_bitfield from " + str(refi) + " for 1) = '1'"
		condstr.append(cstr)
	SQLcond = " (" + " AND ".join(condstr) + ") "

	return SQLcond

CondStrSelectedTripDays = CondStrSelectedTripDays_hex

def CondStrExcludeGattungs(GattungList):
	"""
	Return SQL condition string for excluded Gattungs.
	example: ExcludeGattungList = ['S', 'RE', 'BUS']
	"""
	condstr = []
	for gattung in GattungList:
		pat = "'" + gattung + "'"
		condstr.append(pat)
	SQLcond = "line_category NOT IN " + "(" + ",".join(condstr) + ")"
	return SQLcond

def CondStrSelectedVerwaltungs(VerwaltungList):
	"""
	Return SQL condition string for selected Verwaltungs (Include-Only Verwaltungs).
	example: VerwaltungList = [11, 23, 111]
	"""
	condstr = []
	for vw in VerwaltungList:
		pat = str(vw)
		condstr.append(pat)
	SQLcond = "management IN " + "(" + ",".join(condstr) + ")"
	return SQLcond

def CondStrExcludeVerwaltungs(VerwaltungList):
	"""
	Return SQL condition string for excluded Verwaltungs.
	example: VerwaltungList = [11, 23, 111]
	"""
	condstr = []
	for vw in VerwaltungList:
		pat = str(vw)
		condstr.append(pat)
	SQLcond = "management NOT IN " + "(" + ",".join(condstr) + ")"
	return SQLcond

def CondStrOneOfSelectedTripDays_hex(DayOrdList):
	"""
	Return SQL condition string for selected trip days.
	Selected connections must be available on at least ONE OF the listed days.

	works with hex code (verkehrstage_hexcode)

	DayOrdList: list of ordinal dates
	"""
	# 9.4. String Functions and Operators
	# http://www.postgresql.org/docs/9.4/static/functions-string.html
	# http://stackoverflow.com/questions/17208945/whats-the-easiest-way-to-represent-a-bytea-as-a-single-integer-in-postgresql

	# start day reference
	StartDayRef = FPLAN_BeginDate.toordinal()
	EndDayRef = FPLAN_EndDate.toordinal()

	# convert date string to day count w.r.t. reference day
	DayRefList = []
	for day in DayOrdList:
		ref = day - StartDayRef 

		if ref < 0 or ref > (EndDayRef - StartDayRef):
			raise Exception("All dates in DayList must be between FPLAN START and END days!")
		DayRefList.append(ref)

	# eliminate recurring dates
	DayRefSet = set(DayRefList)

	condstr = []
	for ref in DayRefSet:
		refi = ref + 3
		# select ('x' || verkehrstage_hexcode)::bit(240) 
		cstr = "substring((('x' || trafficdays_hexcode)::bit(" + str(refi) + "))::text from " + str(refi) + " for 1) = '1'"
		condstr.append(cstr)
	SQLcond = " (trafficdays_hexcode is null OR \n(" + "\n OR ".join(condstr) + "))\n "

	# test
	# print "\nSQLcond: +++++++++++++++++"
	# print SQLcond
	return SQLcond

CondStrOneOfSelectedTripDays = CondStrOneOfSelectedTripDays_hex

# **************************************************************************************
# Condition to be applied on ConnectionInfo functions for selecting connections
# **************************************************************************************

def CheckMinStationCount(ConnectionInfo, PathInfo, MinStationCount, EndStation, RouteConditions):
	"""
	Check if the number of stations of route (N) remains less than the lower limit 
	with the next connection. None means no limits.
	"""
	if MinStationCount == None: return True
	NextStation = ConnectionInfo[ConnInfoInd['station_to']]

	if CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation) and len(PathInfo) < (MinStationCount-1):
		return False
	else:
		return True

def CheckMaxStationCount(ConnectionInfo, PathInfo, MaxStationCount):
	"""
	Check if the number of stations of route (N) exceeds the upper limit 
	with the next connection. None means no limits.
	"""
	if MaxStationCount == None: return True

	if len(PathInfo) >= MaxStationCount:
		return False 
	else:
		return True

def CheckMaxWaitTimeAtStation(ConnectionInfo, PathInfo, MaxWaitTime):
	"""
	Check if max waiting time at the station is exceeded. 
	Return false if waiting time is exceeded, otherwise true.
	"""
	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	departure_next_station = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]
	
	if (departure_next_station - arrival_last_station) <= MaxWaitTime: 	
		return True
	else:
		return False

def CheckIfEachStationIsVisitedOnlyOnce(ConnectionInfo, PathInfo, EndStation, RouteConditions):
	"""
	Simple path condition: Each station is visited only once,
	possibly excluding the first and last stations for a circular path.
	"""
	NextStation = ConnectionInfo[ConnInfoInd['station_to']]

	# make an exception for EndStation = FirstStation
	if len(PathInfo) > 1:
		if CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation): 
			return True

	# check if next station was already visited
	StationList = GetAllStationsOfRoute(PathInfo)

	if NextStation not in StationList: 
		return True
	else: 
		return False

def CheckDurationWithNextConnection(ConnectionInfo, PathInfo, LatestArrivalIn):
	"""
	Return true if the total duration of path + connection is 
	less that LatestArrivalIn, otherwise false. None for LatestArrivalIn means no limits.

	Note: 
	Starting ref PathInfo[0] --> LatestArrivalIn determines latest arrival
	Starting ref PathInfo[1] --> LatestArrivalIn determines max duration of route
	"""
	if not PathInfo: return True
	# if len(PathInfo) < 2: return True

	departure_first_station = PathInfo[0][ConnInfoInd['departure_hour']]*60 \
		+ PathInfo[0][ConnInfoInd['departure_min']]

	arrival_next_station = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 \
		+ ConnectionInfo[ConnInfoInd['arrival_min']]

	if LatestArrivalIn == None: LatestArrivalIn = 999999999;
	TotalDuration = arrival_next_station - departure_first_station 

	if TotalDuration <= LatestArrivalIn:
		return True
	else:
		return False

def CheckMinDurationWithNextConnection(ConnectionInfo, PathInfo, EarliestArrivalIn, EndStation, RouteConditions):
	"""
	Check if duration of the whole path is larger than EarliestArrivalIn.
	Return true if:
	a) next station is not EndStation, or 
	b) next station = EndStation and duration >= EarliestArrivalIn, or
	c) EarliestArrivalIn = None. None for EarliestArrivalIn means no limits.

	Note: 
	Starting ref PathInfo[0] --> EarliestArrivalIn determines earliest arrival
	Starting ref PathInfo[1] --> EarliestArrivalIn determines min duration of route
	"""
	if not PathInfo: return True
	# if len(PathInfo) < 2: return True
	if EarliestArrivalIn == None: return True
	if EarliestArrivalIn == 0: return True

	# check next station
	NextStation = ConnectionInfo[ConnInfoInd['station_to']]
	# if NextStation != EndStation: return True
	if not CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation):
		return True

	# get total duration
	departure_first_station = PathInfo[0][ConnInfoInd['departure_hour']]*60 \
		+ PathInfo[0][ConnInfoInd['departure_min']]

	arrival_next_station = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 \
		+ ConnectionInfo[ConnInfoInd['arrival_min']]

	TotalDuration = arrival_next_station - departure_first_station 

	if TotalDuration >= EarliestArrivalIn:
		return True
	else:
		return False

def CheckIfStationIsNotInExcludeList(ConnectionInfo, ExcludedStations):
	"""
	Return false if next station is  in the list of to-be excluded stations; 
	otherwise true.
	ExcludedStations: List of stations to be exluded in the route
	"""
	if ExcludedStations == None or len(ExcludedStations) == 0:
		return True

	NextStation = ConnectionInfo[ConnectionInfoInd['next_station']]

	if NextStation in ExcludedStations:
		return False
	else:
		return True

def CheckIfStationIsInIncludeOnlyList(ConnectionInfo, IncludedStations):
	"""
	Return true if next station is in the list of included stations; otherwise false.
	IncludedStations: List of all stations that should be included in the route
	"""
	if not IncludedStations:
		return False 

	NextStation = ConnectionInfo[ConnInfoInd['station_to']]

	if NextStation in IncludedStations:
		return True
	else:
		return False

def CheckIfMandatoryStationCanBeMeasured(PathInfo, ConnectionInfo, IncludedStations, RequiredMeasureTime):
	"""
	Assumption: Simple Path Route (i.e. every station is visited only once).
	Return false if current station is in the list IncludedStations
	(i.e. must be measured) but there is not enough time for measurement.
	IncludedStations: List of distinct stations that must be measured
	RequiredMeasureTime: Required time in minutes for measuring a station
	"""
	if len(PathInfo) == 0:
		return True

	if not IncludedStations:
		return True 

	CurrStation = ConnectionInfo[ConnectionInfoInd['curr_station']]
	
	# check if current station must be measured
	if CurrStation not in IncludedStations:
		return True

	# time at station
	ankunft_std = PathInfo[-1][ConnInfoInd['arrival_hour']]
	ankunft_min = PathInfo[-1][ConnInfoInd['arrival_min']]
	ankunftm = ankunft_std*60 + ankunft_min

	abfahrt_std = ConnectionInfo[ConnInfoInd['departure_hour']]
	abfahrt_min = ConnectionInfo[ConnInfoInd['departure_min']]
	abfahrtm = abfahrt_std*60 + abfahrt_min

	if (abfahrtm - ankunftm) >= RequiredMeasureTime:
		return True
	else:
		False

def CheckIfMandatoryLineCanBeMeasured(PathInfo, ConnectionInfo, IncludedLines, RequiredMeasureTime):
	"""
	Assumption: Every line (lineID) is visited only once.
	Return false if current line is in the list IncludedStations
	(i.e. must be measured) but there is not enough time for measurement.
	IncludedLines: List of distinct lines that must be measured
	RequiredMeasureTime: Required time in minutes for measuring a station
	"""
	if len(PathInfo) < 2:
		return True

	if not IncludedLines:
		return True 

	# get last line
	LastLine =  PathInfo[-1][ConnInfoInd['line_id']]
	LastTripID = PathInfo[-1][ConnInfoInd['travel_id']]
	
	# check if last line must be measured
	if LastLine not in IncludedLines:
		return True

	# get time spent on last line (in minutes)
	LastArrival = PathInfo[-1][ConnInfoInd['arrival_hour']] * 60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	LastDeparture = PathInfo[-1][ConnInfoInd['departure_hour']] * 60 + PathInfo[-1][ConnInfoInd['departure_min']]

	for i in range(2,len(PathInfo)-1):
		line = PathInfo[-i][ConnInfoInd['line_id']]

		if line == LastLine:
			LastDeparture = PathInfo[-i][ConnInfoInd['departure_hour']] * 60 + PathInfo[-i][ConnInfoInd['departure_min']]
		else:
			break
	TotalTimeOnLine = LastArrival - LastDeparture

	if TotalTimeOnLine >= RequiredMeasureTime:
		return True
	else:
		# get next line
		NextLine = ConnectionInfo[ConnInfoInd['line_id']]
		NextTripID = ConnectionInfo[ConnInfoInd['travel_id']]

		# continue with the same line if TotalTimeOnLine < RequiredMeasureTime
		if NextLine == LastLine and LastTripID == NextTripID:
			return True 
		else:
			return False

def CheckIfEnoughTimeForLineChange(ConnectionInfo, PathInfo, MinChangeTime=5):
	"""
	Check if there is enough time for changing the line.
	A line is uniquely defined by last_fahrtid.
	"""
	if len(PathInfo) <= 1: return True 	# no change time is required at starting station

	# bypass for BU2019 project
	# return MinChangeTime

	# next station
	next_fahrtnum = ConnectionInfo[ConnInfoInd['travel_no']]
	next_abfahrt_std = ConnectionInfo[ConnInfoInd['departure_hour']]
	next_abfahrt_min = ConnectionInfo[ConnInfoInd['departure_min']]
	next_gattung = ConnectionInfo[ConnInfoInd['line_category']]
	next_fahrtid = ConnectionInfo[ConnInfoInd['travel_id']]
	next_line = ConnectionInfo[ConnInfoInd['line_id']]
	
	# last station
	last_fahrtnum = PathInfo[-1][ConnInfoInd['travel_no']]
	last_ankunft_std = PathInfo[-1][ConnInfoInd['arrival_hour']]
	last_ankunft_min = PathInfo[-1][ConnInfoInd['arrival_min']]
	last_gattung = PathInfo[-1][ConnInfoInd['line_category']]
	last_fahrtid = PathInfo[-1][ConnInfoInd['travel_id']]
	last_line = PathInfo[-1][ConnInfoInd['line_id']]

	# total minutes
	NextAbfahrtMin = next_abfahrt_std*60 + next_abfahrt_min
	LastAnkunftMin = last_ankunft_std*60 + last_ankunft_min

	# exception for OnFoot station passages (Haltestellen-Übergang)
	if next_gattung in TrWay.values() or last_gattung in TrWay.values():
		if NextAbfahrtMin >= LastAnkunftMin: return True
	
	# other connections (excluding OnFoot)
	if next_fahrtid == last_fahrtid:
		return True
	else:
		if ( NextAbfahrtMin - LastAnkunftMin) >= MinChangeTime:
			return True
		else:
			return False

def CheckIfConnectionIsAvailableForAllListedDays(ConnectionInfo, DayOrdList):
	"""
	Check if connection is found for all ordinal dates listed in DayOrdList
	"""
	# see:
	# http://stackoverflow.com/questions/2803852/python-date-string-to-date-object
	# http://stackoverflow.com/questions/14524322/how-to-convert-a-date-string-to-different-format

	if not DayList:
		return False 

	# verkehrstage
	VerkehrstageHex = ConnectionInfo[ConnInfoInd['trafficdays_hexcode']]

	# Null or empty VerkehrstageHex --> connection is available for all days
	if not VerkehrstageHex: return True

	VerkehrstageBin = hextobin(VerkehrstageHex)

	# start day reference
	StartDayRef = FPLAN_BeginDate.toordinal()
	EndDayRef = FPLAN_EndDate.toordinal()

	# convert date string to day count w.r.t. reference day
	DayRefList = []
	for day in DayOrdList:
		ref = day - StartDayRef 

		if ref < 0 or ref > (EndDayRef - StartDayRef):
			raise Exception("All dates in DayList must be between FPLAN START and END days!")
		DayRefList.append(ref)

	# eliminate recurring dates
	DayRefSet = set(DayRefList)

	# check all dates
	for ref in DayRefSet:
		if VerkehrstageBin[2 + ref] == '0':
			return False 
	# passed all day ref checks
	return True

def CheckIfEarliestDeparture(ConnectionInfo, PathInfo):
	"""
	Check if the next connection is the earliest departure for a given line, at given station.
	Relevant for shortest-path search to eliminate later depatures for the same line and direction.
	see: http://stackoverflow.com/questions/279561/what-is-the-python-equivalent-of-static-variables-inside-a-function
	"""
	if not hasattr(CheckIfEarliestDeparture, "DepartureTimeDic"):
		CheckIfEarliestDeparture.DepartureTimeDic = {}

	# return True

	curr_station = ConnectionInfo[ConnInfoInd['station_from']]
	next_station = ConnectionInfo[ConnInfoInd['station_to']]

	line_id = ConnectionInfo[ConnInfoInd['line_id']]
	
	NextLine = (curr_station, next_station, line_id)
	departure_next_station = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 + ConnectionInfo[ConnInfoInd['arrival_min']]

	if CheckIfEarliestDeparture.DepartureTimeDic.has_key(next_station):
		earliest_departure = CheckIfEarliestDeparture.DepartureTimeDic[NextLine]

		if departure_next_station < earliest_departure:
			CheckIfEarliestDeparture.DepartureTimeDic[NextLine] = departure_next_station
			print "CheckIfEarliestDeparture: True"
			return True
		else:
			print "CheckIfEarliestDeparture: False"
			return False
	else:
		CheckIfEarliestDeparture.DepartureTimeDic[NextLine] = departure_next_station
		return True

def CheckMaxNumberOfLineChanges(ConnectionInfo, PathInfo, MaxLineChangeLimit):
	"""
	Check if number of line changes <= MaxLineChangeLimit
	A station passage alone is not counted as an additional line change.
	"""
	# first connection from the starting point is not counted as line change
	if len(PathInfo) <= 1: return True 

	# short-cut
	if len(PathInfo) <= MaxLineChangeLimit: return True

	# a station passage is not counted as a line change
	OnFootGattungList = TrWay.values()		# ZF and ZF+M
	next_gattung = ConnectionInfo[ConnInfoInd['line_category']]
	if next_gattung in OnFootGattungList: return True 

	# count line changes 
	LineChanges = 0 
	for i in range(2, len(PathInfo)):
		if PathInfo[i-1][ConnInfoInd['travel_id']] != PathInfo[i][ConnInfoInd['travel_id']]:
			LineChanges += 1
	if ConnectionInfo[ConnInfoInd['travel_id']] != PathInfo[-1][ConnInfoInd['travel_id']]:
		LineChanges += 1

	if LineChanges <= MaxLineChangeLimit:
		return True 
	else:
		return False

def CheckIfLastLineCanBeMeasured(PathInfo,IncludedLines,RequiredMeasureTime):

	# Assumption: Every line (lineID) is visited only once.
	# Return false if current line is in the list IncludedStations
	# (i.e. must be measured) but there is not enough time for measurement.
	# IncludedLines: List of distinct lines that must be measured
	# RequiredMeasureTime: Required time in minutes for measuring a station
	
	# get last line
	LastLine =  PathInfo[-1][ConnInfoInd['line_id']]
	
	# check if last line must be measured
	if LastLine not in IncludedLines:
		return False

	# get time spent on last line (in minutes)
	LastArrival = PathInfo[-1][ConnInfoInd['arrival_hour']] * 60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	LastDeparture = PathInfo[-1][ConnInfoInd['departure_hour']] * 60 + PathInfo[-1][ConnInfoInd['departure_min']]

	for i in range(2,len(PathInfo)-1):
		line = PathInfo[-i][ConnInfoInd['line_id']]

		if line == LastLine:
			LastDeparture = PathInfo[-i][ConnInfoInd['departure_hour']] * 60 + PathInfo[-i][ConnInfoInd['departure_min']]
		else:
			break
	TotalTimeOnLine = LastArrival - LastDeparture

	if TotalTimeOnLine >= RequiredMeasureTime:
		return True
	else:
		return False

def BestStartingStation(RequirementCluster,StationListForLines,EarliestArrival):
	
	RequirementsSet=set(list(list(zip(*RequirementCluster)[0])))
	RequirementsSet=list(RequirementsSet)

	AllRelevantStations=set()
	for line in RequirementsSet:
		AllRelevantStations.update(StationListForLines.loc[line,:].dropna())
	
	AllRelevantStations=list(AllRelevantStations)
	#print AllRelevantStations
	
	TotalDistance=np.zeros(len(AllRelevantStations))
	for station in AllRelevantStations:
		ind=AllRelevantStations.index(station)
		TotalDistance[ind]=EarliestArrival.loc[station,AllRelevantStations].sum(axis=0)
	
	min_index=np.argmin(TotalDistance)
	
	return int(AllRelevantStations[min_index])

def CheckIfReturnIsPossibleFromCurrentStation(ConnectionInfo,PathInfo,RouteConditions,EarliestArrival_StationScoring):
	
	arrival_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	TotalDuration = arrival_last_station - arrival_first_station
	
	if TotalDuration + EarliestArrival_StationScoring.at[0,ConnectionInfo[ConnInfoInd['station_to']]] < RouteConditions[Cond.StartTimeAndDuration][3] :
		return True
	else:
		return False

def GetTimeAndDurationOfPath(PathInfo):
	"""
	Get begin/end times, and total duration of path since arrival to first station (Depot).
	"""
	if not PathInfo:
		return None 
	if len(PathInfo) < 2: return None 

	arrival_first_station = PathInfo[0][ConnInfoInd['arrival_hour']]*60 + PathInfo[0][ConnInfoInd['arrival_min']]
	# departure_first_station = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]

	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	TotalDuration = arrival_last_station - arrival_first_station
	return (TotalDuration, arrival_first_station, arrival_last_station)



def CheckIfPathTerminatesSuccessfully(ConnectionInfo, PathInfo, RouteConditions, EndStation):

	#Check if the route terminates successfully by reaching the end station, 
	#or by reaching one of the stations or LineIDs.

	if not PathInfo or len(PathInfo) < 2:
		return False 

	# get station and LineID in ConnectionInfo
	next_station = ConnectionInfo[ConnInfoInd['station_to']]
	next_line = ConnectionInfo[ConnInfoInd['line_id']]

	if Cond.SuccessfulTerminationBy in RouteConditions:
		(TerminationType, StationOrLineIDList) = RouteConditions[Cond.SuccessfulTerminationBy]
		if TerminationType == Cond.ReachingOneOfTheStations:
			if next_station in StationOrLineIDList:
				return True 
			else:
				return False 
		elif TerminationType == Cond.ReachingOneOfTheLineIDs:
			if next_line in StationOrLineIDList:
				return True 
			else:
				return False 
		else:
			raise Exception("Undefined TerminationType %s!" % TerminationType)

	# standard termination
	
	# Update on 2017.08.23 by Tunc, extra condition for earliest arrival time
	# Update on 2017.10.18 by Tunc, new cond EnableMultipleRoundToursAroundStartingStation for round tours
	EarliestArrivalCond = True

	# bypass for BU2019 project
	"""
	if Cond.EnableMultipleRoundToursAroundStartingStation in RouteConditions and RouteConditions[Cond.EnableMultipleRoundToursAroundStartingStation][0]:
		(StartHour, StartMin, MinTimeSinceStart, MaxTimeSinceStart) = RouteConditions[Cond.StartTimeAndDuration]
		arrival_time = ConnectionInfo[ConnInfoInd['arrival_hour']] * 60 + ConnectionInfo[ConnInfoInd['arrival_min']]
		earliest_arrival_time =  StartHour * 60 + StartMin + MinTimeSinceStart
		EarliestArrivalCond = arrival_time >= earliest_arrival_time
	"""

	if next_station == EndStation and EarliestArrivalCond:
		return True 

	return False

def CheckTripDurationSinceDepartureFromFirstStation(ConnectionInfo, PathInfo, MaxTripDuration):
	"""
	Return TRUE if MaxTripDuration is NOT exceeded since departure from the first station.
	"""
	if len(PathInfo) <= 1: return True

	first_departure = PathInfo[1][ConnInfoInd['departure_hour']]*60 + PathInfo[1][ConnInfoInd['departure_min']]
	next_arrival = ConnectionInfo[ConnInfoInd['arrival_hour']]*60 + ConnectionInfo[ConnInfoInd['arrival_min']]
	TripDuration = next_arrival - first_departure

	if TripDuration <= MaxTripDuration:
		return True 
	else:
		return False

def CheckIfDirectLineConnectionFromStation1ToStation2WasAlreadyTaken(PathInfo, Station1, Station2, EarliestStartFromStation1=None, 
	LatestArrivalToStation1=None, LineAttributeTupleList=None):
	"""
	Check tour (PathInfo) and return True if a direct line connection from Station1 to Station2 is already taken
	(within the given time intervals. Also consider possible constraints LineAttributeTupleList).
	
	Station1: From station
	Station2: To station

	EarliestStartFromStation1: None or a total minute value like 8*60.
		A direct line to Station2 must be taken after EarliestStartFromStation1.
		(EarliestArrivalToStation1 for DirectLine connection from Station1 to Station2)

	LatestArrivalToStation1: Either None, or a total minute value like 9*60.
		Return False if Station1 not reached until LatestArrivalToStation1.
	
	LineAttributeTupleList: Either None or list of line attribute tuples (FahrtID, LineID, fahrtnum, verwaltung, gattung, linie)
		like [(...), (...)]. Departure line must match one of these tuples. 
		attribute=None means (like FahrtID=None) don't consider attribute for matching.
		Not None: Take only a direct line to Station2, that matches one of the listed attribute tuples. 

	Created on 26.12.2017 by Tunc A. Kütükcüoglu
	"""
	# shortcut
	if not PathInfo: return False 
	if len(PathInfo) < 2: return False 

	StationChain = GetAllStationsOfRoute(PathInfo)
	if not Station1 in StationChain or not Station2 in StationChain:
		return False 

	# Station1 and Station2 are both in StationChain, continue...

	# get index of Station1 in PathInfo
	ind1 = None
	for i in range(0, len(PathInfo)):
		if EarliestStartFromStation1 != None:
			if PathInfo[i][ConnInfoInd['arrival_hour']]*60 + PathInfo[i][ConnInfoInd['arrival_min']] < EarliestStartFromStation1:
				continue
		if LatestArrivalToStation1 != None:
			if PathInfo[i][ConnInfoInd['arrival_hour']]*60 + PathInfo[i][ConnInfoInd['arrival_min']] > LatestArrivalToStation1:
				break
		if PathInfo[i][ConnInfoInd['station_to']] == Station1:
			ind1 = i
			break

	# return False if Station2 is not visited after Station1
	ind2 = None
	for i in range(ind1, len(PathInfo)):
		if PathInfo[i][ConnInfoInd['station_to']] == Station2:
			ind2 = i 
			break 
	if not ind2:
		return False

	# check if single (and correct) line from Station1 to Station2
	FahrtID = None
	
	for i in range(ind1+1, ind2+1):
		Conn = PathInfo[i]
		fahrt_id = Conn[ConnInfoInd['travel_id']]
		
		if i == ind1+1:
			line_id = Conn[ConnInfoInd['line_id']]
			fahrtnum = Conn[ConnInfoInd['travel_no']]
			verwaltung = Conn[ConnInfoInd['management']]
			gattung = Conn[ConnInfoInd['line_category']]
			linie = Conn[ConnInfoInd['line']]
			FahrtID = fahrt_id

			LineMatchFound = False

			if not LineAttributeTupleList:
				LineMatchFound = True 
			else:
				for AttributeTuple in LineAttributeTupleList:
					if AttributeTuple[0] and AttributeTuple[0] != FahrtID:
						continue 
					if AttributeTuple[1] and AttributeTuple[1] != line_id:
						continue 
					if AttributeTuple[2] and AttributeTuple[2] != fahrtnum:
						continue 
					if AttributeTuple[3] and AttributeTuple[3] != verwaltung:
						continue 
					if AttributeTuple[4] and AttributeTuple[4] != gattung:
						continue 
					if AttributeTuple[5] and AttributeTuple[5] != linie:
						continue 
					LineMatchFound = True 
					break

			# no match
			if not LineMatchFound: return False

		if i > ind1+1:
			if fahrt_id != FahrtID:
				return False 

	# single line checks passed 
	return True

def CheckIfOnFootPassageFromStation1ToStation2WasAlreadyTaken(PathInfo, Station1, Station2,
	EarliestStartFromStation1=None, LatestArrivalToStation1=None, MaxPassageCount=2):
	"""
	Check tour (PathInfo) and return True if an OnFoot passage from Station1 to Station2 is already taken
	(within the given time intervals).

	Station1: From station
	Station2: To station

	EarliestStartFromStation1: None or a total minute value like 8*60.
		An OnFoot connection to Station2 must be taken after EarliestStartFromStation1.
		(EarliestArrivalToStation1 for OnFoot passage from Station1 to Station2)

	LatestArrivalToStation1: Either None, or a total minute value like 9*60.
		Return False if Station1 not reached until LatestArrivalToStation1.

	Created on 26.12.2017 by Tunc A. Kütükcüoglu
	"""
	# shortcut
	if not PathInfo: return False 
	if len(PathInfo) < 2: return False 

	StationChain = GetAllStationsOfRoute(PathInfo)
	if not Station1 in StationChain or not Station2 in StationChain:
		return False 

	# Station1 and Station2 are both in StationChain, continue...

	# get index of Station1 in PathInfo
	ind1 = None
	for i in range(0, len(PathInfo)):
		if EarliestStartFromStation1 != None:
			if PathInfo[i][ConnInfoInd['arrival_hour']]*60 + PathInfo[i][ConnInfoInd['arrival_min']] < EarliestStartFromStation1:
				continue
		if LatestArrivalToStation1 != None:
			if PathInfo[i][ConnInfoInd['arrival_hour']]*60 + PathInfo[i][ConnInfoInd['arrival_min']] > LatestArrivalToStation1:
				break
		if PathInfo[i][ConnInfoInd['station_to']] == Station1:
			ind1 = i
			break

	# return False if Station2 is not visited after Station1
	ind2 = None
	for i in range(ind1, len(PathInfo)):
		if PathInfo[i][ConnInfoInd['station_to']] == Station2:
			ind2 = i 
			break 
	if not ind2:
		return False

	# check if direct OnFoot passage(s) from Station1 to Station2
	OnFootGattungs = TrWay.values() 

	for i in range(ind1+1, ind2+1):
		Conn = PathInfo[i]
		gattung = Conn[ConnInfoInd['line_category']]
		if not gattung in OnFootGattungs:
			return False 

	# all checks passed
	return True

def CheckIfStationPairsAreConnectedWithASingleLine(ConnectionInfo, PathInfo, RouteConditions, ListOfStationPairs):
	"""
	Check if station pairs ListOfStationPairs are connected with a single line, in given order,
	without line changes between the stations in a pair.

	Builds AggregatePath based on ListOfStationPairs and calls CheckIfAggregatePathIsFollowedRelatively
	"""
	# build AggregatePath
	AggregatePath = []
	for (station1, station2) in ListOfStationPairs:
		AggrConn = CreateAggregateConnection(haltestelle_ab=station1, haltestelle_an=station2)
		AggregatePath.append(AggrConn)
	
	return CheckIfAggregatePathIsFollowedRelatively(ConnectionInfo, PathInfo, RouteConditions, AggregatePath)

def CheckIfStationsAreVisitedInGivenOrder(ConnectionInfo, PathInfo, RouteConditions, OrderedStationList):
	"""
	Check if stations are visited in the order given in OrderedStationList.

	In a station list like [S1, S2, S3, ...] S2 cannot be visited before S1.
	Or generally, Sn cannot be visited before Sm, if n > m.

	If the last visited station is Sn, return false if next station is included 
	in OrderedStationList before Sn.

	An already visited station in OrderedStationList can be visited again.
	"""
	# shortcuts
	if not PathInfo or len(PathInfo) < 2:
		return True 
	if not OrderedStationList or len(OrderedStationList) < 2:
		return True 

	# return true if next station is not in OrderedStationList
	NextStation = ConnectionInfo[ConnInfoInd['station_to']]
	if not NextStation in OrderedStationList:
		return True 
	else:
		# get last (highest-order) already visited station in OrderedStationList
		LastListedStation = None
		MaxInd = -1
		for i in range(1, len(PathInfo)+1):
			station = PathInfo[-i][ConnInfoInd['station_to']]
			
			if station in OrderedStationList:
				ind = OrderedStationList.index(station)
				if ind > MaxInd:
					LastListedStation = station 
					MaxInd = ind

		# check station orders (an equal or lower order station can be visited again)
		NextStationIND = OrderedStationList.index(NextStation) + 1

		LastStationIND = 0
		if LastListedStation:
			LastStationIND = OrderedStationList.index(LastListedStation) + 1

		if NextStationIND <= LastStationIND + 1:
			return True 
		else:
			return False

def CheckMinimumWaitingTimeAtGivenStations(ConnectionInfo, PathInfo, RouteConditions, MinWaitingTimePerStation):
	"""
	Check if line stops (waits) at given stations. Minimum waiting time is indicated for each station.

	This condition is valid only for the first time when a given station is reached.
	If the same station is visited more than once, the waiting condition does not apply 
	for the 2. and later visits.

	1) Return True if LastStation (of path) is not in MinWaitingTimePerStation
	2) Return True if LastStation (of path) is in MinWaitingTimePerStation, 
			but it is not visited for the first time 
	3) Return False, if LastStation (of path) is in MinWaitingTimePerStation, 
			and it is visited for the first time, 
			and NextDeparture - LastArrival < MinWaitingTime 
	"""
	LastStation = PathInfo[-1][ConnInfoInd['station_to']]
	if not LastStation in MinWaitingTimePerStation:
		return True 

	# check if LastStation was visited before 
	for i in range(2, len(PathInfo) + 1):
		station = PathInfo[-i][ConnInfoInd['station_to']]
		if station == LastStation:
			return True 

	# check time difference 
	LastArrival = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
	NextDeparture = ConnectionInfo[ConnInfoInd['departure_hour']]*60 + ConnectionInfo[ConnInfoInd['departure_min']]

	if NextDeparture - LastArrival < MinWaitingTimePerStation[LastStation]:
		return False  
	else:
		return True

# **************************************************************************************
# Condition functions to be applied on PathInfo for selecting routes
# **************************************************************************************

def CheckIfRouteIncludesAllStationsInList(PathInfo, IncludedStations):
	"""
	Return true if route contains all the stations in the list IncludedStations;
	otherwise false.
	"""
	if not IncludedStations:
		return True 

	path = set(GetAllStationsOfRoute(PathInfo))
	stations = set(IncludedStations)

	if path.issuperset(stations):
		return True 
	else:
		return False

def CheckIfRouteExludesAllStationInList(PathInfo, ExcludedStations):
	"""
	Return true if route does not contain any station 
	listed in ExcludedStations; otherwise false.
	"""
	if ExcludedStations == None or len(ExcludedStations) == 0:
		return True 

	path = set(GetAllStationsOfRoute(PathInfo))
	stations = set(IncludedStations)

	if len(path.intersection(stations)) == 0:
		return True 
	else:
		return False

def CheckIfAllLinesCanBeMeasured(PathInfo, TimeIntervalsForLineMeasurement, TimeRequiredForMeasurement):
	"""
	Return true if all the lines in dictionary TimeIntervalsForLineMeasurement 
	can be measured within the given time intervals; otherwise false.

	TimeIntervalsForLineMeasurement[line] = (TimeIntBegin, TimeIntEnd); time limits in minutes
	TimeRequiredForMeasurement: Required time for a line measurement in minutes
	"""
	AllLinesOfPath = GetAllLinesOfPath(PathInfo)
	AllLinesToMeasure = TimeIntervalsForLineMeasurement.keys()

	if not AllLinesToMeasure:
		return True
	if not AllLinesOfPath:
		return False

	TimeIntervalOnLine = GetTimeIntervalsForEachLine(PathInfo)

	for Line in AllLinesToMeasure:
		MeasureIntervalBegin = TimeIntervalsForLineMeasurement[Line][0]
		MeasureIntervalEnd = TimeIntervalsForLineMeasurement[Line][1]

		TimeIntvM = GetTimeIntervalForMeasurementOnLine(Line, TimeIntervalOnLine, MeasureIntervalBegin, 
			MeasureIntervalEnd, TimeRequiredForMeasurement)
		if TimeIntvM == None:
			return False
	return True

def CheckIfSomeOfTheLinesCanBeMeasured(PathInfo, LineList, TimeIntervalForLineMeasurement, TimeRequiredForMeasurement):
	"""
	Return True if at least one of the lines in LineList can be measured
	within the given time interval.

	LineList = [LineID1, LineID2, ...]
	TimeIntervalForLineMeasurement = (560, 600), in minutes 
	TimeRequiredForMeasurement: Required time for a line measurement in minutes
	"""
	AllLinesOfPath = GetAllLinesOfPath(PathInfo)

	if not LineList:
		return True
	if not AllLinesOfPath:
		return False

	TimeIntervalOnLine = GetTimeIntervalsForEachLine(PathInfo)
	MeasureIntervalBegin = TimeIntervalForLineMeasurement[0]
	MeasureIntervalEnd = TimeIntervalForLineMeasurement[1]

	for Line in LineList:	
		TimeIntvM = GetTimeIntervalForMeasurementOnLine(Line, TimeIntervalOnLine, MeasureIntervalBegin, 
			MeasureIntervalEnd, TimeRequiredForMeasurement)
		if TimeIntvM:
			return True
	return False

def CheckIfAllStationsCanBeMeasured(PathInfo, TimeIntervalsForStationMeasurement, TimeRequiredForMeasurement):
	"""
	Return true if all the stations in dictionary TimeIntervalsForStationMeasurement 
	can be measured within the given time intervals; otherwise false.

	TimeIntervalsForStationMeasurement[station] = (TimeIntBegin, TimeIntEnd); time limits in minutes
	TimeRequiredForMeasurement: Required time for a station measurement in minutes
	"""
	AllStationsOfPath = GetAllStationsOfRoute(PathInfo)
	AllStationsToMeasure = TimeIntervalsForStationMeasurement.keys()

	if not AllStationsToMeasure:
		return True
	if not AllStationsOfPath:
		return False

	for Station in AllStationsToMeasure:
		TimeIntervalAtStation = GetTimeIntervalsForEachStation(PathInfo)
		MeasureIntervalBegin = TimeIntervalsForStationMeasurement[Station][0]
		MeasureIntervalEnd = TimeIntervalsForStationMeasurement[Station][1]

		TimeIntvM = GetTimeIntervalForMeasurementAtStation(Station, TimeIntervalAtStation, MeasureIntervalBegin, 
			MeasureIntervalEnd, TimeRequiredForMeasurement)
		if TimeIntvM == None:
			return False
	return True

def CheckIfEarliestArrivalRouteSoFar(PathInfo, CheckMinLineChange):
	"""
	Route selection for earliest arrival.

	Return false if CheckMinLineChange = True and MinLineChangeCount (achieved so far) is exceeded,
	or false/none if minumum duration (achieved so far) is exceeded.
	"""
	if len(PathInfo) <= 1: return True

	if CheckMinLineChange:	
		# count line changes 
		LineChanges = GetNumberOfLineChanges(PathInfo, ConnectionInfo=None)

		if Cond.MinLineChangeCount == None or LineChanges <= Cond.MinLineChangeCount:
			Cond.MinLineChangeCount = LineChanges
		else:
			return False

	# get arrival time
	arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]

	if Cond.EarliestArrival == None or arrival_last_station <= Cond.EarliestArrival:
		Cond.EarliestArrival = arrival_last_station
		
		# test
		# print "arrival_last_station = %s" % arrival_last_station
		
		return True
	else:
		return False

def CheckIfPathIsAvailableSomeOfListedDays(PathInfo, DayList):
	"""
	Return True if route is availabe on at least one of the days listed 
	in DayList (ordinal dates).
	"""
	(AvailableDaysRoute, UnavailableDaysRoute) = GetAvailabilityOfRoute(PathInfo, PeriodBegin, PeriodEnd)
	
	AvailableDaysRoute = set(AvailableDaysRoute)
	DaySet = set(DayList)
	
	if AvailableDaysRoute.intersection(DaySet):
		return True 
	else:
		return False

# **************************************************************************************
# Filter functions to be applied on PathInfoList for selecting or sorting paths
# **************************************************************************************

def FindShortestPath(PathInfoList):
	"""
	Find shortest path in the path list for earliest arrival to destination
	"""
	min_time = 999999999;
	min_time_index = None

	for i in range(0, len(PathInfoList)):
		PathInfo = PathInfoList[i]
		arival_time = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
		if arival_time < min_time:
			min_time_index = i
			min_time = arival_time	

	# return path list with the shortest path only
	ShortestPathInfo = PathInfoList[min_time_index]
	return [ShortestPathInfo]

def SelectRoutesForEarliestArrival(PathInfoList, CheckMinLineChange):
	"""
	Select route(s) for earliest arrival.
	Note: multiple routes can have the same earliest arrival time.
	"""
	if not PathInfoList:
		return PathInfoList

	SelectedRoutes = []
	for PathInfo in PathInfoList:
		# check line changes
		if CheckMinLineChange:
			LineChanges = GetNumberOfLineChanges(PathInfo, ConnectionInfo=None)
			if LineChanges > Cond.MinLineChangeCount:
				continue 
		
		# check arrival time
		arrival_last_station = PathInfo[-1][ConnInfoInd['arrival_hour']]*60 + PathInfo[-1][ConnInfoInd['arrival_min']]
		if arrival_last_station <= Cond.EarliestArrival:
			SelectedRoutes.append(PathInfo)
	return SelectedRoutes

# **************************************************************************************
# Route Operations
# **************************************************************************************

def AddRouteBtoRouteA(RouteA, RouteB, MaxWaitingTime, DefaultLineChangeTime, LineChangeTimePerStation):
	"""
	Add RouteB to the tail of RouteA. 

	1) The departure station of RouteB must be identical to the final arrival station of RouteA.
	2) The time difference between final arrival of RouteA, and first departure of RouteB
		must be equal to or smaller than MaxWaitingTime.
	"""
	if not RouteA:
		return RouteB

	ArrivalTimeA = RouteA[-1][ConnInfoInd['ankunft_std']]*60 + RouteA[-1][ConnInfoInd['ankunft_min']]
	DepartureTimeB = RouteB[1][ConnInfoInd['abfahrt_std']]*60 + RouteB[1][ConnInfoInd['abfahrt_min']]

	ArrivalStationA = RouteA[-1][ConnInfoInd['next_station']]
	DepartureStationB = RouteB[1][ConnInfoInd['curr_station']]

	TripIDA = RouteA[-1][ConnInfoInd['trip_id']]
	TripIDB = RouteB[1][ConnInfoInd['trip_id']]

	if ArrivalStationA != DepartureStationB:
		# raise Exception("First departure station of RouteB must be identical to final arrival station of RouteA.")
		print "First departure station of RouteB must be identical to final arrival station of RouteA."
		return None 
	
	if DepartureTimeB < ArrivalTimeA:
		# raise Exception("First departure of RouteB can not be earlier than final arrival of RouteA.")
		print "First departure of RouteB can not be earlier than final arrival of RouteA."
		return None

	if DepartureTimeB - ArrivalTimeA > MaxWaitingTime:
		# raise Exception("DepartureTimeB - ArrivalTimeA > MaxWaitingTime")
		print "DepartureTimeB - ArrivalTimeA > MaxWaitingTime"
		return None

	LineChangeTime = DefaultLineChangeTime
	if ArrivalStationA in LineChangeTimePerStation:
		LineChangeTime = LineChangeTimePerStation[ArrivalStationA] 

	if TripIDA != TripIDB and DepartureTimeB - ArrivalTimeA < LineChangeTime:
		# raise Exception("DepartureTimeB - ArrivalTimeA < LineChangeTime")
		print "DepartureTimeB - ArrivalTimeA < LineChangeTime"
		return None

	return RouteA + RouteB[1:]

def ApplyAllRouteInfoCorrections(RouteInfo):
	"""
	Apply all corrections to RouteInfo here for missing or incorrect data.

	1) Non-real tunnel stations like 138 (StationNr < 1000). LineID assignment may also be missing
		for such connections.
	"""
	# correction for (1)
	CorrectedRouteInfo = []
	PrevLineID = None 
	PrevFahrtID = None
	HstNrLimit = 1000
	CorrectedConnInfo = None 
	IfCorrectNextConn = False

	# check connnection by connection
	for ConnInfo in RouteInfo:
		haltestelle_ab = ConnInfo[ConnInfoInd['station_from']]
		haltestelle_an = ConnInfo[ConnInfoInd['station_to']]
		CurrLineID = ConnInfo[ConnInfoInd['line_id']]
		ankunft_std = ConnInfo[ConnInfoInd['arrival_hour']]
		ankunft_min = ConnInfo[ConnInfoInd['arrival_min']]
		fahrt_id = ConnInfo[ConnInfoInd['travel_id']]

		if haltestelle_an < HstNrLimit:
			CorrectedConnInfo = list(ConnInfo)
			IfCorrectNextConn = True 
			continue

		if haltestelle_ab < HstNrLimit:
			if ConnInfo[ConnInfoInd['line_id']]:
				CorrectedConnInfo[ConnInfoInd['line_id']] = ConnInfo[ConnInfoInd['line_id']]
			else:
				if fahrt_id == PrevFahrtID and PrevLineID:
					CorrectedConnInfo[ConnInfoInd['line_id']] = PrevLineID

			CorrectedConnInfo[ConnInfoInd['arrival_hour']] = ankunft_std
			CorrectedConnInfo[ConnInfoInd['arrival_min']] = ankunft_min
			CorrectedConnInfo[ConnInfoInd['station_to']] = haltestelle_an 
			CorrectedRouteInfo.append(CorrectedConnInfo)
			IfCorrectNextConn = False 
			continue

		CorrectedRouteInfo.append(ConnInfo)
		PrevLineID = CurrLineID
		PrevFahrtID = fahrt_id

	# return corrected RouteInfo
	return CorrectedRouteInfo


# test module
if __name__ == '__main__':

	dbcon = psycopg2.connect(**PrimaryDB) 
	dbcur = dbcon.cursor()
