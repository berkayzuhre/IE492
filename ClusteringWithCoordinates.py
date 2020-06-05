import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from matplotlib import pyplot as plt
from openpyxl import Workbook
import openpyxl
from GlobalVariables import RequirementScores

line_coordinates=pd.read_excel('coordinates.xlsx')
line_coordinates=line_coordinates.set_index('Line',drop=True)

line_coordinates_numpy=line_coordinates.to_numpy(copy=True)
wcss = []
for i in range(1, 20):
    kmeans = KMeans(n_clusters=i, init='k-means++', max_iter=300, n_init=10, random_state=0)
    #print np.any(np.isnan(RequirementScores_numpy))
    #print np.all(np.isfinite(RequirementScores_numpy))
    kmeans.fit(line_coordinates_numpy)
    wcss.append(kmeans.inertia_)

plt.plot(range(1, 20), wcss)
plt.title('Elbow Method')
plt.xlabel('Number of clusters')
plt.ylabel('WCSS')
plt.show()

ChosenNumberOfClusters=4

kmeans=KMeans(n_clusters=ChosenNumberOfClusters, init='k-means++', max_iter=300, n_init=10, random_state=0)
kmeans.fit(line_coordinates_numpy)
workbook = openpyxl.Workbook()
sheet = workbook.active

for i in range(1,len(kmeans.labels_)):

	sheet.cell(row=i, column=1, value=RequirementScores.columns.values[i-1])
	sheet.cell(row=i, column=2, value=kmeans.labels_[i-1])

workbook.save(filename="clusters_with_coordinates.xlsx")